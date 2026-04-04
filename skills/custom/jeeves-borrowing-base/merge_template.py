#!/usr/bin/env python3
"""Merge data tabs from a borrowing base workbook into a template.

Downloads the latest template from Google Drive, replaces data tabs with
fresh data, preserves formula/summary tabs, and saves the result.

Usage:
    python merge_template.py <data_workbook> <template_drive_id> [--output <path>]

The script auto-detects which tabs to replace based on the data workbook's
sheet names (tape_start, tape_end, rollforward, eligibility_summary, tape_combined).
"""

import os
import re
import sys
import tempfile
from urllib.parse import parse_qs, urlparse


def _ensure_deps():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-q', 'openpyxl'])


def _extract_id(url_or_id):
    if re.match(r'^[a-zA-Z0-9_-]+$', url_or_id):
        return url_or_id
    m = re.search(r'/d/([a-zA-Z0-9_-]+)', url_or_id)
    if m:
        return m.group(1)
    parsed = urlparse(url_or_id)
    qs = parse_qs(parsed.query)
    if 'id' in qs:
        return qs['id'][0]
    raise ValueError(f"Cannot extract file ID from: {url_or_id}")


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(SCRIPT_DIR, '..', '_shared'))
from google_auth import get_credentials


def _download_template(drive_id):
    """Download an xlsx from Drive to a temp file."""
    from googleapiclient.discovery import build

    file_id = _extract_id(drive_id)
    creds = get_credentials()
    service = build('drive', 'v3', credentials=creds)

    meta = service.files().get(fileId=file_id, fields='name,mimeType').execute()
    name = meta.get('name', 'template')
    print(f"Downloading template: {name}", file=sys.stderr)

    content = service.files().get_media(fileId=file_id).execute()
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', prefix='template_', delete=False)
    tmp.write(content)
    tmp.close()
    return tmp.name, name


def _clear_sheet(ws):
    """Clear all cell values in a worksheet without deleting it."""
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None


def _copy_sheet_data(source_ws, target_ws, *, add_index: bool = False):
    """Copy cell values from one worksheet to another.

    Args:
        add_index: If True, insert a numeric index in column A and shift
            data one column to the right.  The original Bridge/SOFOM
            templates have a pandas index column (col A = 0, 1, 2...)
            that formula tabs reference.  ``build_us.py`` writes data
            with ``index=False``, so we re-add it here.
    """
    col_offset = 1 if add_index else 0

    for row in source_ws.iter_rows():
        for cell in row:
            target_col = cell.column + col_offset
            target_cell = target_ws.cell(row=cell.row, column=target_col)
            target_cell.value = cell.value
            if cell.number_format:
                target_cell.number_format = cell.number_format

    if add_index:
        # Write index values in column A (header = blank, rows = 0, 1, 2...)
        # Skip row 1 (header row — leave blank like pandas does)
        for r in range(2, source_ws.max_row + 1):
            target_ws.cell(row=r, column=1, value=r - 2)

    # Copy column widths (shifted if index was added)
    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width


def main():
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')

    import argparse
    parser = argparse.ArgumentParser(description='Merge data into template')
    parser.add_argument('data_workbook', help='Local path to data workbook')
    parser.add_argument('template_drive_id', help='Drive ID or URL of template')
    parser.add_argument('--output', default=None, help='Output path (default: $OUTPUTS_PATH)')
    args = parser.parse_args()

    _ensure_deps()
    import openpyxl

    if not os.path.isfile(args.data_workbook):
        print(f"ERROR: Data workbook not found: {args.data_workbook}", file=sys.stderr)
        sys.exit(1)

    # Download template
    template_path, template_name = _download_template(args.template_drive_id)

    try:
        # Open both workbooks
        print("Opening template (preserving formulas)...")
        template_wb = openpyxl.load_workbook(template_path)
        original_tab_order = list(template_wb.sheetnames)  # capture before any modifications

        print("Opening data workbook...")
        data_wb = openpyxl.load_workbook(args.data_workbook)

        data_sheet_names = data_wb.sheetnames
        print(f"Data tabs to merge: {data_sheet_names}")

        # Track which template sheets are formula tabs (not in data workbook)
        formula_tabs = [s for s in template_wb.sheetnames if s not in data_sheet_names]
        print(f"Template formula tabs (preserved): {formula_tabs}")

        # Tabs that need a pandas index column (col A) to match the
        # original template structure.  elig_summary does not have one.
        # NOTE: 'tape' (SOFOM) and 'tape_combined' do NOT get an index -
        # the SOFOM template's tape tab starts with 'dt' in col A.
        _TABS_WITH_INDEX = {'tape_start', 'tape_end', 'rollforward'}

        # Replace data tabs in-place (clear + rewrite) to preserve tab order.
        # If a data tab doesn't exist in the template yet, append it.
        for name in data_sheet_names:
            source_ws = data_wb[name]
            needs_index = name in _TABS_WITH_INDEX

            if name in template_wb.sheetnames:
                print(f"  Replacing '{name}' in template (in-place)...")
                target_ws = template_wb[name]
                _clear_sheet(target_ws)
            else:
                print(f"  Adding new '{name}' to template...")
                target_ws = template_wb.create_sheet(name)

            _copy_sheet_data(source_ws, target_ws, add_index=needs_index)

        # Determine output path
        if args.output:
            output_path = args.output
        else:
            output_dir = os.environ.get('OUTPUTS_PATH', '/mnt/user-data/outputs')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, template_name)

        # Restore original template tab order -- openpyxl shifts sheets to the front
        # when clearing/rewriting them in-place. Re-apply the order captured before
        # any modifications. Any sheets not in original_tab_order are appended at end.
        sheet_map = {ws.title: ws for ws in template_wb._sheets}
        ordered = [n for n in original_tab_order if n in sheet_map]
        extras = [n for n in sheet_map if n not in original_tab_order]
        try:
            template_wb._sheets = [sheet_map[n] for n in ordered + extras]
        except Exception:
            pass

        print(f"Saving merged workbook: {output_path}")
        template_wb.save(output_path)
        print(f"\nDone! Output: {output_path}")
        print("Open in Excel to let formulas recalculate.")

    finally:
        # Cleanup temp file
        if os.path.exists(template_path):
            os.unlink(template_path)


if __name__ == '__main__':
    main()
