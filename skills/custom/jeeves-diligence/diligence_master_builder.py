import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.environ.get('OUTPUTS_PATH', '/mnt/user-data/outputs')
os.makedirs(OUT, exist_ok=True)
path = os.path.join(OUT, 'Diligence Master Tracker - All Workstreams - 20260615.xlsx')

# ---- styling ----
HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
SEC_FILL = PatternFill('solid', fgColor='D9E1F2')
SEC_FONT = Font(bold=True, color='1F3864', size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F3864')
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='top')
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

STATUS_FILL = {
    'Complete': PatternFill('solid', fgColor='C6EFCE'),
    'In Progress': PatternFill('solid', fgColor='FFEB9C'),
    'Open': PatternFill('solid', fgColor='FFC7CE'),
    'Confirm': PatternFill('solid', fgColor='FCE4D6'),
    'N/A': PatternFill('solid', fgColor='F2F2F2'),
}
STATUS_FONT = {
    'Complete': Font(color='006100', bold=True),
    'In Progress': Font(color='9C6500', bold=True),
    'Open': Font(color='9C0006', bold=True),
    'Confirm': Font(color='974706', bold=True),
    'N/A': Font(color='808080'),
}

COLS = ['Ref', 'Category', 'Request', 'Owner', 'Status', 'Notes / Evidence', 'Link']
WIDTHS = [10, 22, 58, 18, 13, 60, 30]

def style_sheet(ws, title, subtitle, rows):
    ws.sheet_view.showGridLines = False
    ws.merge_cells('A1:G1'); ws['A1'] = title; ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:G2'); ws['A2'] = subtitle
    ws['A2'].font = Font(italic=True, size=9, color='595959')
    for i, w in enumerate(WIDTHS):
        ws.column_dimensions[get_column_letter(i+1)].width = w
    r = 4
    for c, h in enumerate(COLS):
        cell = ws.cell(r, c+1, h); cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CENTER; cell.border = BORDER
    ws.freeze_panes = f'A{r+1}'
    r += 1
    for row in rows:
        if row.get('section'):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            cell = ws.cell(r, 1, row['section']); cell.fill = SEC_FILL; cell.font = SEC_FONT
            cell.alignment = Alignment(vertical='center'); cell.border = BORDER
            ws.row_dimensions[r].height = 18
            r += 1; continue
        vals = [row.get('ref',''), row.get('cat',''), row.get('req',''),
                row.get('owner',''), row.get('status',''), row.get('notes',''), row.get('link','')]
        for c, v in enumerate(vals):
            cell = ws.cell(r, c+1, v); cell.border = BORDER; cell.alignment = WRAP
        st = row.get('status','')
        scell = ws.cell(r, 5)
        if st in STATUS_FILL:
            scell.fill = STATUS_FILL[st]; scell.font = STATUS_FONT[st]; scell.alignment = CENTER
        ws.cell(r, 1).alignment = CENTER
        ws.row_dimensions[r].height = 42
        r += 1
    return ws

wb = Workbook()

# ===================== TAB 0: SUMMARY =====================
ws0 = wb.active; ws0.title = 'Summary'
ws0.sheet_view.showGridLines = False
for i, w in enumerate([34, 16, 14, 16, 64]):
    ws0.column_dimensions[get_column_letter(i+1)].width = w
ws0.merge_cells('A1:E1'); ws0['A1'] = 'Jeeves — Diligence Master Tracker'; ws0['A1'].font = TITLE_FONT
ws0.merge_cells('A2:E2'); ws0['A2'] = 'Consolidated across all live workstreams · As of 2026-06-15 · Maintained by DeerFlow-Analyst'
ws0['A2'].font = Font(italic=True, size=9, color='595959')
hdr = ['Workstream', 'Counterparty', 'Items', 'Open/IP', 'Current state']
r = 4
for c, h in enumerate(hdr):
    cell = ws0.cell(r, c+1, h); cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
summary = [
    ['1. VCP + PwC FDD (financial)', 'Vista Credit', 18, '2 confirm', 'Effectively closed — all 18 delivered to VDR. Residual confirms: VCP-4 PN memo; FDD-7/8 completeness.'],
    ['2. Legal DD request list (PSK)', 'Vista (Proskauer)', 38, '38 open', 'New legal list (v5, 6/13). Sections 1–13 incl. large Brazilian DD. Needs triage + owners.'],
    ['3. Legal/Credit DD (Akin)', 'Neuberger Berman (Akin Gump)', 15, '15 open', 'Loan/receivables list (12) + HR/compliance gaps (3). Upload folders empty. Tax/audit via Goodwin.'],
    ['4. Collateral Audit (Kroll)', 'Kroll (engaged by NB)', 16, '4 HAVE / 5 PARTIAL / 7 NEED', 'SOW 1 only. 10 Phase 1 + 6 Phase 2 (P2 pending sample). Fee $65–75K (vs ~$40K target). Kick-off Wed 6/18 11:30am ET.'],
]
r = 5
for row in summary:
    for c, v in enumerate(row):
        cell = ws0.cell(r, c+1, v); cell.border = BORDER; cell.alignment = WRAP if c==4 else CENTER
        if c == 0: cell.alignment = Alignment(vertical='center'); cell.font = Font(bold=True)
    ws0.row_dimensions[r].height = 40
    r += 1
ws0.cell(r+1, 1, 'Legend:').font = Font(bold=True)
leg = [('Complete','Complete'),('In Progress','In progress / partial'),('Open','Open / not started'),('Confirm','Delivered – needs confirmation')]
for i,(k,lab) in enumerate(leg):
    cell = ws0.cell(r+2+i, 1, lab); cell.fill = STATUS_FILL[k]; cell.font = STATUS_FONT[k]; cell.alignment = CENTER

# ===================== TAB 1: VISTA VCP + PwC FDD =====================
DDF = 'https://drive.google.com/drive/folders/1JiMVbmEKQxKN_e5TXY2MWY3LXJGqv48P'
vista = [
 {'section':'VCP DATA REQUESTS — Emily Dang (VCP), reconciled vs. VDR 2026-06-15'},
 {'ref':'VCP-1','cat':'VCP Data','req':'Unblinded full customer-level ledger (TPV + revenue per customer, geo + product tags)','owner':'Brian','status':'Complete','notes':'Customer ledger as of Apr 30 2026; 6,634 accounts, $161.2M.','link':'VCP-1 Customer Ledger Data Tape'},
 {'ref':'VCP-2','cat':'VCP Data','req':'Sales pipeline by geo/product/stage/probability + historical conversion & quota attainment','owner':'Josh Pudnos','status':'Complete','notes':'Equity VDR files verbatim (Pipeline Analytics + GTM Quota vs Attainment).','link':'VCP-2 Pipeline + Quota'},
 {'ref':'VCP-3','cat':'VCP Data','req':'Flash YTD May 2026 results (TPV, revenue, gross profit, EBITDA)','owner':'Pradeep','status':'Complete','notes':'Delivered 6/11 — "Jeeves May 2026 Prelim P&L". (Updated since stale 6/10 tracker.)','link':'VCP-3 May 2026 Prelim P&L'},
 {'ref':'VCP-4','cat':'VCP Data','req':'Historical stablecoin volume by country & provider + license/regulatory confirmation','owner':'Brandon / Pradeep','status':'Confirm','notes':'Reg summary PDF + monthly TPV table (delivered 6/14) both in VDR. CONFIRM: separate PN regulatory memo (Brandon) complete.','link':'VCP-4 Reg Summary + Stablecoin TPV'},
 {'ref':'VCP-5','cat':'VCP Data','req':'Bridge/Circle custodial account statements + master custodial agreement','owner':'Brian','status':'Complete','notes':'In VDR.','link':DDF},
 {'ref':'VCP-6','cat':'VCP Data','req':'AIG trade credit insurance utilization (claims, losses vs sublimits, capacity)','owner':'Brian','status':'Complete','notes':'Summary PDF: policy bound early 2026; no claims (180 DPD threshold); capacity undrawn.','link':'VCP-6 AIG Utilization'},
 {'section':'PwC FDD DATA REQUESTS — Financial Due Diligence, reconciled vs. VDR 2026-06-15'},
 {'ref':'FDD-1','cat':'PwC FDD','req':'Monthly consolidating IS + BS trial balances FY25/LTM26 by country, with account-to-FS mapping','owner':'Shalom','status':'Complete','notes':'Monthly P&L FY2025 + Quarterly B/S FY2025. 2026 periods removed per Alex.','link':'FDD-1 P&L + B/S'},
 {'ref':'FDD-2','cat':'PwC FDD','req':'FY24 (and if avail FY25) audited financials, rep letters, control observations, misstatements','owner':'Alex','status':'Complete','notes':'FY2023 + FY2024 audited financials in Diligence Library. FY25 not yet audited.','link':'FDD-2 Audited Financials'},
 {'ref':'FDD-3','cat':'PwC FDD','req':'Monthly revenue breakdown by stream/product/country/entity/currency','owner':'Pradeep','status':'Complete','notes':'FDD Responses 6/10 — tabs FDD3-MX/CO/BR/RoW.','link':'FDD-3-4-5 Revenue/Take Rate/NII'},
 {'ref':'FDD-4','cat':'PwC FDD','req':'Monthly gross & net take rate by product/country, reconciled to mgmt reporting + GL','owner':'Pradeep','status':'Complete','notes':'FDD Responses 6/10 — tab FDD4 LATAM Take Rates & GP.','link':'FDD-3-4-5 Revenue/Take Rate/NII'},
 {'ref':'FDD-5','cat':'PwC FDD','req':'Monthly NII margin analysis FY25/LTM26 (yield, funding costs, NII, provisions, hedging)','owner':'Pradeep','status':'Complete','notes':'FDD Responses 6/10 — tab FDD5 LATAM UE Extended.','link':'FDD-3-4-5 Revenue/Take Rate/NII'},
 {'ref':'FDD-6','cat':'PwC FDD','req':'Flow of funds diagrams + SOPs (interchange, Jeeves/Instant Pay, lending) by country/entity','owner':'Isa','status':'Complete','notes':'Vista cut: LATAM cards + Jeeves Pay, 10 slides (Isa FinOps deck Apr 2026).','link':'FDD-6 Flow of Funds'},
 {'ref':'FDD-7','cat':'PwC FDD','req':'Settlement timing/lag, float balances, working capital, settlement receivables/payables recon','owner':'Shalom / Pradeep','status':'Confirm','notes':'Deliverable in VDR 6/12 ("Net Corporate Cash" / cash proxy). CONFIRM it satisfies settlement-timing/float ask (was Request Sent on stale tracker).','link':'FDD-7-8 Net Corporate Cash'},
 {'ref':'FDD-8','cat':'PwC FDD','req':'Monthly cash bridge by entity + full bank account listing (cash traps, reserves, repatriation)','owner':'Shalom / Pradeep','status':'Confirm','notes':'Cash proxy delivered 6/12. CONFIRM full bank-account listing by entity/currency is included (NB/Akin also asking — see Akin tab).','link':'FDD-7-8 Net Corporate Cash'},
 {'ref':'FDD-9','cat':'PwC FDD','req':'Updated lending portfolio rollforward + cash movement through latest FY26 date','owner':'Brian','status':'Complete','notes':'LOC Acct Rollforward to Apr 30 2026, 162,533 rows.','link':'FDD-9 LOC Acct Rollforward'},
 {'ref':'FDD-10','cat':'PwC FDD','req':'Monthly ALLL rollforward FY25/LTM26 by product/country + DQ/NPL/CO + reserve methodology','owner':'Shalom','status':'Complete','notes':'DQ+CO by country (Brian) + ALLL/CECL rollforward delivered 6/12 ("Allowance for Loan Loss Roll" sheet). Shalom piece now in.','link':'FDD-10 CECL rollforward'},
 {'ref':'FDD-11','cat':'PwC FDD','req':'Summary of all outstanding debt facilities + Dec-25 MX ABS facility detail','owner':'Brian','status':'Complete','notes':'Facility Overview VDR; Bridge rate corrected to SOFR + 645–925 bps.','link':'FDD-11 Facility Overview'},
 {'ref':'FDD-12','cat':'PwC FDD','req':'FX accounting & hedging support (functional currency, methodology, gains/losses, hedging)','owner':'Shalom','status':'Complete','notes':'Jeeves FX Accounting Policy Memo — Final, delivered 6/9.','link':'FDD-12 FX Accounting Memo'},
]
style_sheet(wb.create_sheet('1 · Vista VCP+PwC'),
            'Workstream 1 — Vista Credit: VCP + PwC FDD (Financial Diligence)',
            'Source: VCP+PwC DDQ Tracker, reconciled against actual VDR folder contents 2026-06-15. DD Package folder linked per row.',
            vista)

# ===================== TAB 2: VISTA LEGAL (PROSKAUER) =====================
psk = [
 {'section':'Standard legal sections (1–9) — Proskauer "Jeeves Information & Document Review Request List", v5 (PR Draft 6/13/26)'},
 {'ref':'1.01','cat':'Parties & Properties','req':'Any change in identity / corporate structure (mergers, consolidations, acquisitions) past 5 yrs','owner':'Legal','status':'Open','notes':''},
 {'ref':'1.02','cat':'Parties & Properties','req':'Address of chief executive office and chief places of business','owner':'Legal','status':'Open','notes':''},
 {'ref':'1.03','cat':'Parties & Properties','req':'List of US & foreign trademarks / trade names / service marks (reg #, date, applications)','owner':'Legal','status':'Open','notes':''},
 {'ref':'1.04','cat':'Parties & Properties','req':'List of US & foreign patents and patent applications','owner':'Legal','status':'Open','notes':''},
 {'ref':'1.05','cat':'Parties & Properties','req':'List of US & foreign copyrights and pending registrations','owner':'Legal','status':'Open','notes':''},
 {'ref':'2.01','cat':'Org & Capitalization','req':'Any DD memos/summaries produced previously or for current transaction','owner':'Legal','status':'Open','notes':''},
 {'ref':'2.02','cat':'Org & Capitalization','req':'Structure chart with TIN and ownership of each Group Company','owner':'Legal','status':'Open','notes':''},
 {'ref':'2.03','cat':'Org & Capitalization','req':'List of all intercompany / junior debt instruments (payor, payee, amount)','owner':'Finance/Legal','status':'Open','notes':''},
 {'ref':'2.04','cat':'Org & Capitalization','req':'List of all directors and officers of each Group Company','owner':'Legal','status':'Open','notes':''},
 {'ref':'3.01','cat':'Contracts & Commitments','req':'All loan agreements, financing arrangements, intercreditor agreements, guaranties','owner':'Brian/Legal','status':'Open','notes':'Maps to FDD-11 Facility Overview — reuse, don\u2019t rebuild.'},
 {'ref':'3.02','cat':'Contracts & Commitments','req':'Top 10 customer & vendor contracts material to the business','owner':'Legal','status':'Open','notes':''},
 {'ref':'3.03','cat':'Contracts & Commitments','req':'Any lien searches conducted re: any Group Company','owner':'Legal','status':'Open','notes':''},
 {'ref':'4.01','cat':'Litigation','req':'All pending/threatened litigation, claims, regulatory investigations (past 3 yrs)','owner':'Legal','status':'Open','notes':''},
 {'ref':'5.01','cat':'Insurance','req':'All insurance policies (coverage, limits, deductibles, expiration, premiums) + copies','owner':'OPS/Legal','status':'Open','notes':''},
 {'ref':'6.01','cat':'Authorizations & Consents','req':'All required governmental / shareholder / 3rd-party consents for the transaction','owner':'Legal','status':'Open','notes':''},
 {'ref':'6.02','cat':'Authorizations & Consents','req':'All permits, licenses, approvals, certifications required to operate','owner':'Legal','status':'Open','notes':'Partial overlap with VCP-4 reg summary.'},
 {'ref':'7.01','cat':'Affiliate Transactions','req':'Intercompany agreements & agreements with equity holders (tax sharing, loans, side letters)','owner':'Legal/Finance','status':'Open','notes':''},
 {'ref':'8.01','cat':'Financial Information','req':'Audited financials (3 yrs) + interim unaudited since most recent audit','owner':'Alex/Finance','status':'Open','notes':'Maps to FDD-2 (FY23/FY24 audited) — reuse.'},
 {'ref':'9.01','cat':'Employee Benefits','req':'Material employee benefit, pension, labor, employment, CBA & related regulatory matters','owner':'HR','status':'Open','notes':''},
 {'section':'10. FCPA & UK Bribery Act'},
 {'ref':'10.01','cat':'FCPA / Bribery','req':'List of clients/customers/partners operating outside the US, and countries','owner':'Compliance','status':'Open','notes':''},
 {'ref':'10.02','cat':'FCPA / Bribery','req':'List of all countries Jeeves has done / does / plans to do business (last 5 yrs)','owner':'Compliance','status':'Open','notes':''},
 {'ref':'10.03','cat':'FCPA / Bribery','req':'List of subsidiaries/agents/subcontractors assisting non-US business + countries','owner':'Compliance','status':'Open','notes':''},
 {'ref':'10.04','cat':'FCPA / Bribery','req':'Anti-bribery policies currently in place','owner':'Compliance','status':'Open','notes':''},
 {'ref':'10.05','cat':'FCPA / Bribery','req':'Schedule of foreign government licenses, permits, approvals','owner':'Compliance/Legal','status':'Open','notes':''},
 {'ref':'10.06','cat':'FCPA / Bribery','req':'Overview of the Company\u2019s compliance program, policies and procedures','owner':'Compliance','status':'Open','notes':''},
 {'section':'11. Cybersecurity, IP & Data Privacy'},
 {'ref':'11.01','cat':'Cyber / IP / Privacy','req':'Measures to maintain secrecy of trade secrets / confidential & customer info','owner':'Security/Legal','status':'Open','notes':''},
 {'ref':'11.02','cat':'Cyber / IP / Privacy','req':'Confirm proprietary tech developed in-house vs. independent contractors','owner':'Eng/Legal','status':'Open','notes':''},
 {'ref':'11.03','cat':'Cyber / IP / Privacy','req':'Confirm NDA + IP assignment agreements (present-tense assignment) for all staff/contractors','owner':'Legal/HR','status':'Open','notes':''},
 {'ref':'11.04','cat':'Cyber / IP / Privacy','req':'Confirm any actual/threatened IP litigation, disputes or proceedings','owner':'Legal','status':'Open','notes':''},
 {'ref':'11.05','cat':'Cyber / IP / Privacy','req':'3rd-party cyber risk assessments + most recent audit + high-risk areas','owner':'Security','status':'Open','notes':''},
 {'ref':'11.06','cat':'Cyber / IP / Privacy','req':'Confirm collection/storage of PII (customers/employees) and health data','owner':'Security/Legal','status':'Open','notes':''},
 {'ref':'11.07','cat':'Cyber / IP / Privacy','req':'Does the Company "sell"/"share" personal data with third parties (per privacy laws)','owner':'Legal','status':'Open','notes':''},
 {'ref':'11.08','cat':'Cyber / IP / Privacy','req':'Confirm compliance with applicable privacy laws, standards & posted privacy policy','owner':'Legal','status':'Open','notes':''},
 {'ref':'11.09','cat':'Cyber / IP / Privacy','req':'Security incidents / data breaches history + investigations/enforcement + exposure','owner':'Security/Legal','status':'Open','notes':''},
 {'ref':'11.10','cat':'Cyber / IP / Privacy','req':'How Company uses AI/ML + internal AI/ML policies (incl. generative AI by staff/vendors)','owner':'Eng/Legal','status':'Open','notes':''},
 {'section':'12. Sustainability'},
 {'ref':'12.01','cat':'Sustainability','req':'Sustainability documentation (sponsor questionnaire responses, policies, KPIs, EDCI)','owner':'Finance/Legal','status':'Open','notes':''},
 {'ref':'12.02','cat':'Sustainability','req':'In-scope assessment for sustainability regs (CSRD, CA climate acts) + compliance steps','owner':'Legal','status':'Open','notes':''},
 {'ref':'12.03','cat':'Sustainability','req':'Formal processes/policies to identify & manage sustainability risks (+ top 5 risks if yes)','owner':'Legal','status':'Open','notes':''},
 {'section':'13. Brazilian Due Diligence (regulator-driven — Card Services, Payments, FX, Virtual Assets/Stablecoins, Regulatory)'},
 {'ref':'13.01','cat':'BR — Card Services','req':'Card issuance & structure: issuer entity, postpaid-issuer licensing, program architecture, credit/financing flow (SCD), payment-instrument acceptance, interface with prepaid payment accounts (5 sub-items)','owner':'Legal BR / Brandon','status':'Open','notes':'Jeeves Brasil (CNPJ 44.302.146/0001-67).'},
 {'ref':'13.02','cat':'BR — Program Agmt','req':'Program Agreement: issuer/acceptance scope; account structure & ownership of funds (individualized vs omnibus) (4 sub-items)','owner':'Legal BR','status':'Open','notes':''},
 {'ref':'13.03','cat':'BR — Jeeves Cash','req':'Jeeves Cash Services (Anexo D): account manager, fund location/ownership, account nature','owner':'Legal BR','status':'Open','notes':''},
 {'ref':'13.04','cat':'BR — FX','req':'Foreign exchange: contracting/settlement flow, eFX services, Bexs FX correspondent agreement (3 sub-items)','owner':'Legal BR / Treasury','status':'Open','notes':''},
 {'ref':'13.05','cat':'BR — QI SCD','req':'Payment Account Partnership w/ QI SCD: BaaS model confirm + related documents','owner':'Legal BR','status':'Open','notes':''},
 {'ref':'13.06','cat':'BR — Virtual Assets','req':'Stablecoins/crypto: VASP intent (BCB Res 519 / Instr 704/26), wallet custody, BRL→stablecoin flow, Bridge & Infinia roles, issuer/backing/jurisdiction, AML/KYC allocation (21 sub-items)','owner':'Legal BR / Brandon','status':'Open','notes':'Largest sub-section. Overlaps VCP-4 reg summary (Bridge).'},
 {'ref':'13.07','cat':'BR — General','req':'SCD/IP authorization status & filings; BCB requests; group policy sharing; regulatory inquiries/incidents; other licenses (7 sub-items)','owner':'Legal BR','status':'Open','notes':''},
 {'ref':'13.08','cat':'BR — General','req':'Internal policies/manuals/procedures (compliance, AML, risk, governance, controls)','owner':'Compliance BR','status':'Open','notes':''},
]
style_sheet(wb.create_sheet('2 · Vista Legal (PSK)'),
            'Workstream 2 — Vista Credit: Legal DD Request List (Proskauer / PSK)',
            'Source: "Vista_Jeeves – Diligence Request List (PR Draft 6/13/26)" v5 + Litera redline vs v4. NEW — needs triage & owner assignment. Brazilian DD sub-items summarized at section level.',
            psk)

# ===================== TAB 3: NB / AKIN =====================
NBF = 'https://drive.google.com/drive/folders/1yae9sldBwAvj3G8fRYY2EdGtE33VPKfX'
akin = [
 {'section':'Akin Gump (NB counsel) — "Preliminary Legal DD Document Request List: Loan/Receivable Contracts & Policies" (contacts: Fernandez / Gilligan / Hawkins @akingump.com)'},
 {'ref':'A-1','cat':'Loan/Receivable','req':'Current underwriting, origination & servicing policies & procedures (incl. prior-period versions)','owner':'Credit Team','status':'In Progress','notes':'Jorge requested 6/15 — data-room copy is partial. Upload → CREDIT folder (empty).','link':NBF},
 {'ref':'A-2','cat':'Loan/Receivable','req':'25 randomly selected electronic loan/receivable files for outstanding loans (representative across geos)','owner':'Credit Team','status':'In Progress','notes':'Jorge requested 6/15. Export from loan management system. Upload → CREDIT folder (empty).','link':NBF},
 {'ref':'A-3','cat':'Loan/Receivable','req':'Survey of (a) interest-rate limits & licensing, (b) commercial loan broker reqs, (c) choice-of-law validity, per jurisdiction','owner':'Legal / Goodwin','status':'Open','notes':'Likely Goodwin-prepared legal survey.'},
 {'ref':'A-4','cat':'Loan/Receivable','req':'List of 10 largest distribution partners by volume + related agreements & fees','owner':'OPS (Victor/Cami)','status':'In Progress','notes':'Jorge requested 6/15 — need structured ranked list. Upload → OPS folder (empty).','link':NBF},
 {'ref':'A-5','cat':'Loan/Receivable','req':'Standard form agreements with partners using the origination platform','owner':'Legal / OPS','status':'Open','notes':''},
 {'ref':'A-6','cat':'Loan/Receivable','req':'Standard form agreements with vendors under the vendor program','owner':'Legal / OPS','status':'Open','notes':''},
 {'ref':'A-7','cat':'Loan/Receivable','req':'Standard form agreements with referral / marketing partners (compensation arrangements)','owner':'Legal / OPS','status':'Open','notes':''},
 {'ref':'A-8','cat':'Loan/Receivable','req':'All form loan & receivable agreements & related docs (current + prior periods)','owner':'Legal','status':'Open','notes':''},
 {'ref':'A-9','cat':'Loan/Receivable','req':'Agreements with credit/bank partners re: origination & sale of loans/receivables','owner':'Legal','status':'Open','notes':''},
 {'ref':'A-10','cat':'Loan/Receivable','req':'Material leases for computer hardware/software/equipment used to originate or service','owner':'OPS','status':'Open','notes':''},
 {'ref':'A-11','cat':'Loan/Receivable','req':'All insurance policies covering charge-card receivables (+ excess/umbrella)','owner':'OPS / Finance','status':'Open','notes':'Related to AIG policy (see Vista VCP-6).'},
 {'ref':'A-12','cat':'Loan/Receivable','req':'All commercial notes generated re: loans outstanding in Brazil','owner':'Legal BR','status':'Open','notes':''},
 {'section':'NB HR / Compliance gap items (Jorge → Audrey, 6/15) — part of broader NB legal DD'},
 {'ref':'A-HR1','cat':'HR / Compliance','req':'Written confirmation: no material disputes/disagreements with external auditors','owner':'Audrey (HR)','status':'Open','notes':'One-line confirmation. Upload → HR folder (empty).','link':NBF},
 {'ref':'A-HR2','cat':'HR / Compliance','req':'Written confirmation: no internal reviews/investigations at request of Board/committee/internal audit/GC','owner':'Audrey (HR)','status':'Open','notes':'One-line confirmation.','link':NBF},
 {'ref':'A-HR3','cat':'HR / Compliance','req':'Complete current whistleblower policy','owner':'Audrey (HR)','status':'Open','notes':''},
 {'section':'NB Tax / Audit / Compliance / Internal Controls (Goodwin → Shalom/Alex, 6/12) — tracked separately'},
 {'ref':'A-TAX','cat':'Tax / Audit','req':'Neuberger diligence questions re: Taxes, Auditing, Compliance & Internal Controls','owner':'Goodwin / Brian / Shalom','status':'In Progress','notes':'Per Brian 6/15: "Goodwin has most of this, I\u2019ll handle what they don\u2019t." Item-level list with Goodwin.'},
 {'section':'Logistics — NB + Akin + Goodwin all-hands: Thursday 6/18, 3:00pm ET'},
]
style_sheet(wb.create_sheet('3 · NB-Akin'),
            'Workstream 3 — Neuberger Berman: Legal/Credit DD (Akin Gump, NB counsel)',
            'Sources: Akin "Preliminary Legal DD Request List" (loan/receivables, 12 items) + Jorge gap distribution 6/15 + NB Tax/Audit thread. Upload folders (CREDIT/HR/OPS) currently EMPTY.',
            akin)

# ===================== TAB 4: KROLL COLLATERAL AUDIT =====================
KF = 'https://drive.google.com/drive/folders/1amlOqm423CkTVYWqukx9HmXE5kTCZiZC'
kroll = [
 {'section':'Engagement / commercial — SOW 1 Collateral Audit (SOW 2&3 removed). Fee $65–75K (revised SOW 6/12, down from ~$85K; Jeeves target ~$40K) + 5% + MX City travel. Prelim 15 biz days from initial data; draft +5. Kick-off Wed 6/18 11:30am ET (Weber MD / Kee SVP / Pauzano / Chrysostomou).'},
 {'section':'PHASE 1 — initial data request (Nick Pauzano list, 15-Jun-2026) · status: 4 HAVE · 5 PARTIAL · 1 NEED'},
 {'ref':'P1-1','cat':'Phase 1','req':'Business overview — org structure, products/services, key management','owner':'CapMkts','status':'Complete','notes':'HAVE: Business Overview Deck (NB 4/22) + Jeeves Cards & Payments Infrastructure deck. Confirm standalone org chart/bios if Kroll wants.','link':KF},
 {'ref':'P1-2','cat':'Phase 1','req':'NB financing agreement + underwriting & eligibility criteria + legal docs (LOI/purchase, loan, servicing)','owner':'CapMkts / Legal','status':'In Progress','notes':'PARTIAL: Credit Agreement (6/1 draft), Executed Term Sheet (4/13), Collateral Proposal. GAP: CA still in negotiation (Goodwin issues list 6/14); eligibility schedule + servicing agreement to confirm.','link':KF},
 {'ref':'P1-3','cat':'Phase 1','req':'Credit scoring, risk grading, collections, servicing, accounting & risk mgmt policies + internal audit/monitoring reports','owner':'Credit / Risk','status':'In Progress','notes':'PARTIAL: Jeeves Credit Policy - Apr 2026 (added) + NB ODD Session 1 deck. GAP: servicing/collections/accounting policies + internal audit/monitoring reports. (Same policies as Akin A-1.)','link':KF},
 {'ref':'P1-4','cat':'Phase 1','req':'Audited financials (2–3 yrs), latest management accounts, business plans, investor decks, portfolio reports','owner':'Finance','status':'In Progress','notes':'PARTIAL: Conservative Forecast Outlook (5/26). GAP: 2024 audited (in Vista FDD pkg), latest management accounts, investor deck, monthly portfolio report.','link':KF},
 {'ref':'P1-5','cat':'Phase 1','req':'Documentation/walkthroughs of key processes — origination, underwriting, funding, servicing, collections, month-end close','owner':'Ops / Finance','status':'In Progress','notes':'PARTIAL: NB ODD Session 1 deck. GAP: process walkthrough write-ups / flow docs per process; month-end close narrative.','link':KF},
 {'ref':'P1-6','cat':'Phase 1','req':'Summary of identified breaches or exceptions to documented policies & procedures','owner':'Credit / Risk','status':'Open','notes':'NEED: exceptions/breach log to be compiled.','link':KF},
 {'ref':'P1-7','cat':'Phase 1','req':'List & details of vendor disputes, litigation, administrative proceedings, or governmental inquiries','owner':'Legal','status':'Complete','notes':'HAVE: "Summary - No Material Litigation" delivered 6/15. CONFIRM it also covers vendor disputes + administrative/governmental inquiries.','link':KF},
 {'ref':'P1-8','cat':'Phase 1','req':'Complete population of NB receivables (inception to date) — master file for sample selection','owner':'CapMkts / Data','status':'Complete','notes':'HAVE: Data Tape - NB Receivables 6/14 (standard one-day tape_end, 6,735 rows). Provide additional dates if Kroll wants a range.','link':KF},
 {'ref':'P1-9','cat':'Phase 1','req':'Customer base summary — key counterparties, concentration metrics, merchant/processor relationships','owner':'CapMkts / Data','status':'Complete','notes':'HAVE: NB Customer Base Concentration (top obligors, tiers, by country/product; active book $171.0M / 2,484 accts). Merchant/processor detail to add if required.','link':KF},
 {'ref':'P1-10','cat':'Phase 1','req':'Overview of payment flows — bank accounts and collection account structure','owner':'Finance / Treasury','status':'In Progress','notes':'PARTIAL: Flow of Funds - FinOps (added) + Cards & Payments deck. GAP: explicit bank-account list + collection-account structure (ties to Akin OPS bank schedule).','link':KF},
 {'section':'PHASE 2 — supporting docs (pending sample selection) · all NEED'},
 {'ref':'P2-1','cat':'Phase 2','req':'Supporting docs for selected receivable samples (applications, executed loan agreements, invoices, funding records)','owner':'Ops / CapMkts','status':'Open','notes':'NEED — pending sample selection.','link':KF},
 {'ref':'P2-2','cat':'Phase 2','req':'Detailed servicing records & payment histories for sampled receivables','owner':'Ops / Data','status':'Open','notes':'NEED — pending sample selection.','link':KF},
 {'ref':'P2-3','cat':'Phase 2','req':'Docs for delinquent/NPL/modified/high-risk accounts in sample — collections actions & recovery efforts','owner':'Credit / Collections','status':'Open','notes':'NEED — pending sample selection.','link':KF},
 {'ref':'P2-4','cat':'Phase 2','req':'Payment tracing support — linking borrower payments to servicing systems, bank records, financial reporting','owner':'Finance / Ops','status':'Open','notes':'NEED — pending sample selection. (Settlement tracing = primary fee driver.)','link':KF},
 {'ref':'P2-5','cat':'Phase 2','req':'Bank statements, cash receipt records & reconciliations linking servicing, bank activity & financial reporting','owner':'Finance / Treasury','status':'Open','notes':'NEED — pending sample selection.','link':KF},
 {'ref':'P2-6','cat':'Phase 2','req':'Internal credit memos & approval docs for sampled borrowers (esp. larger/higher-risk exposures)','owner':'Credit','status':'Open','notes':'NEED — pending sample selection. (Kroll requests credit memos for top 10 by exposure.)','link':KF},
]
style_sheet(wb.create_sheet('4 · Kroll'),
            'Workstream 4 — Kroll Collateral Audit (NB Receivables; Kroll engaged by NB)',
            'Source: "Tracker - Kroll Collateral Audit Requests - 20260615.xlsx" (Nick Pauzano request list) in NB Diligence/Kroll folder. 16 items (10 P1 / 6 P2). Statuses reconciled vs actual Kroll folder files 2026-06-15 17:47 (P1-7 promoted to HAVE; P1-3/P1-10 material refreshed).',
            kroll)

wb.save(path)
print('SAVED', path)
