#!/usr/bin/env python3
"""cap_markets_metrics_refresh.py
=================================
Pull fresh Capital Markets metrics from Drive (BB files) and Redshift
(portfolio strats), write to .deer-flow/_bb_metrics_state.json, then call
cap_markets_metrics_writer.py to merge cron-health into _cap_markets_state.json
so the gateway /metrics endpoint exposes everything to Grafana.

Designed to run standalone from the cron supervisor — no HTML template
dependency, no Drive upload. Reuses the BB-parsing functions from
skills/custom/jeeves-capital-markets/dashboard_full_update.py.

Idempotent. Safe to run on a 1-hour cadence. If Redshift is unreachable
(Zscaler down), BB/SOFOM still populate via Drive; portfolio panels stay
on their previous values.
"""
from __future__ import annotations

import json
import logging
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv

BACKEND_DIR = Path(__file__).resolve().parent.parent
REPO_DIR = BACKEND_DIR.parent.parent
SKILLS_DIR = REPO_DIR / "deer-flow" / "skills"
STATE_DIR = BACKEND_DIR / ".deer-flow"
OUT_FILE = STATE_DIR / "_bb_metrics_state.json"

# Load .env so GOOGLE_*, REDSHIFT_* are available when run outside the gateway.
load_dotenv(BACKEND_DIR / ".env")

# Allow imports from the BB skill (for redshift_util) and the capital-markets skill
sys.path.insert(0, str(SKILLS_DIR / "custom" / "jeeves-borrowing-base"))
sys.path.insert(0, str(SKILLS_DIR / "custom" / "jeeves-capital-markets"))

# OUTPUTS / WORKSPACE need to exist before importing dashboard_full_update —
# the module uses them at import time. Point them at a scratch dir under
# .deer-flow so the import succeeds; we never actually write HTML here.
SCRATCH = STATE_DIR / "_cap_markets_scratch"
SCRATCH.mkdir(parents=True, exist_ok=True)
os.environ.setdefault("SKILLS_PATH", str(SKILLS_DIR))
os.environ.setdefault("OUTPUTS_PATH", str(SCRATCH))
os.environ.setdefault("WORKSPACE_PATH", str(SCRATCH))

# Reuse the existing parsing logic. dashboard_full_update is a script-style
# module so the heavy imports happen at module load; that's acceptable here.
import dashboard_full_update as dfu  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("cap-markets-refresh")


# ISO 3166-1 numeric country code → label. Add codes as needed.
_COUNTRY_CODES = {
    32: "Argentina", 40: "Austria", 76: "Brazil", 124: "Canada",
    152: "Chile", 170: "Colombia", 233: "Estonia", 276: "Germany",
    380: "Italy", 484: "Mexico", 528: "Netherlands", 604: "Peru",
    620: "Portugal", 724: "Spain", 826: "United Kingdom", 840: "United States",
}


def _country_label(code) -> str:
    if isinstance(code, (int, float)):
        return _COUNTRY_CODES.get(int(code), f"Other ({int(code)})")
    if isinstance(code, str) and code.strip():
        return code.strip()
    return "Unknown"


def _dpd_bucket(dpd) -> str:
    if not isinstance(dpd, (int, float)):
        return "Current"
    d = int(dpd)
    if d <= 0:
        return "Current"
    if d <= 30:
        return "1-30 DPD"
    if d <= 60:
        return "31-60 DPD"
    if d <= 90:
        return "61-90 DPD"
    return "90+ DPD"


def _download_xlsx(svc, file_id: str):
    """Download .xlsx from Drive, open in Excel via xlwings to force a
    full recalc (so formula cells get cached values), then re-load with
    openpyxl in data_only mode.

    Why: BB workbooks are produced programmatically (pandas/xlsxwriter) so
    they have formulas but no cached values. openpyxl's data_only=True
    returns None for those cells. xlwings shells out to Excel COM, forces
    a full recalc, and saves — populating the cached values openpyxl
    needs. This is the only way to get the canonical BB numbers without
    reimplementing every Excel formula in Python (the workbooks use
    Office-365 LET/UNIQUE/FILTER which no pure-Python evaluator supports).
    """
    import io
    import tempfile
    from googleapiclient.http import MediaIoBaseDownload
    from openpyxl import load_workbook

    req = svc.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    dl = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = dl.next_chunk()

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(fh.getvalue())
        tmp_path = tmp.name

    try:
        import xlwings as xw
        app = xw.App(visible=False, add_book=False)
        try:
            wb_xl = app.books.open(tmp_path)
            wb_xl.app.calculate()
            wb_xl.save()
            wb_xl.close()
        finally:
            app.quit()
    except Exception as e:
        log.warning("xlwings recalc failed (%s) — falling back to formula-only read", e)

    wb = load_workbook(tmp_path, data_only=True)
    try:
        os.unlink(tmp_path)
    except OSError:
        pass
    return wb


def _read_us_bridge(wb) -> dict:
    """Extract US Bridge BB metrics — canonical values from the Summary
    sheet when available (after xlwings recalc populates the cached cells),
    falling back to tape-derived estimates.

    Summary sheet cells (USD-denominated):
      L40  Total Eligible Receivables
      L60  Total Receivables Counted Towards BB (post-concentration)
      L94  Receivable BB component (= L60)
      L95  US Cash BB component
      L96  ex-US Cash BB component
      L97  Mastercard Liability Amount (negative — deduction)
      L99  Receivable Advance Rate
      L100 US Cash Advance Rate
      L101 ex-US Cash Advance Rate
      L102 Mastercard Liability Advance Rate
      L104 Borrowing Base = SUMPRODUCT(L94:L97, L99:L102)
      L106 Total Drawn
      L107 Advance (additional)
      L108 BB Excess / Deficit (= L104 - L106 - L107)
    """
    out: dict = {}

    if "Summary" in wb.sheetnames:
        sm = wb["Summary"]
        # Build a label → row map by scanning column B. The Bridge BB
        # template's row numbers shift periodically as concentration tests
        # are added/renumbered (saw a shift on 2026-06-07), so resolve every
        # canonical cell by its B-column label instead of hardcoded row.
        b_label_to_row: dict[str, int] = {}
        for row in range(1, sm.max_row + 1):
            v = sm[f"B{row}"].value
            if isinstance(v, str):
                stripped = v.strip()
                b_label_to_row.setdefault(stripped, row)

        def _by_label(label: str, col: str = "L"):
            r = b_label_to_row.get(label)
            if r is None:
                return None
            v = sm[f"{col}{r}"].value
            return float(v) if isinstance(v, (int, float)) else None

        eligible_total = _by_label("Total Eligible Receivables")
        receivables_counted = _by_label("Total Receivables Counted Towards Borrowing Base")
        us_cash_bb = _by_label("US Cash Borrowing Base")
        exus_cash_bb = _by_label("ex-US Cash Borrowing Base")
        mc_liability = _by_label("Mastercard Liability Amount")
        rcv_rate = _by_label("Receivable Advance Rate")
        usc_rate = _by_label("US Cash Advance Rate")
        exusc_rate = _by_label("ex-US Cash Advance Rate")
        mc_rate = _by_label("Mastercard Liability Advance Rate")
        bb = _by_label("Borrowing Base")
        drawn = _by_label("Total Drawn")
        advance = _by_label("Advance")
        excess_deficit = _by_label("Borrowing Base Excess / Deficit")
        facility_size = _by_label("Facility Size")

        if eligible_total is not None:       out["eligible_gross"] = round(eligible_total, 2)
        if receivables_counted is not None:  out["eligible"] = round(receivables_counted, 2)
        if eligible_total is not None and receivables_counted is not None:
            out["concentration_excess_total"] = round(eligible_total - receivables_counted, 2)
        if us_cash_bb is not None:           out["us_cash"] = round(us_cash_bb, 2)
        if exus_cash_bb is not None:         out["ex_us_cash"] = round(exus_cash_bb, 2)
        if mc_liability is not None:         out["mastercard_liability"] = round(mc_liability, 2)
        if rcv_rate is not None:             out["receivable_advance_rate"] = rcv_rate
        if usc_rate is not None:             out["us_cash_advance_rate"] = usc_rate
        if exusc_rate is not None:           out["ex_us_cash_advance_rate"] = exusc_rate
        if mc_rate is not None:              out["mastercard_advance_rate"] = mc_rate
        if bb is not None:                   out["borrowing_base"] = round(bb, 2)
        if drawn is not None:                out["total_drawn"] = round(drawn, 2)
        if advance is not None:              out["additional_advance"] = round(advance, 2)
        if excess_deficit is not None:       out["bb_excess_deficit"] = round(excess_deficit, 2)
        # The file's Facility Size cell reads $50M (template default) but
        # the actual CIM Bridge facility has been upsized — hardcoded.
        # If/when the template gets refreshed, remove this override.
        BRIDGE_FACILITY_SIZE_OVERRIDE = 75_000_000.0
        out["facility_size"] = BRIDGE_FACILITY_SIZE_OVERRIDE
        if facility_size is not None:        out["facility_size_template"] = round(facility_size, 2)

        # Bridge concentration covenants — find the "Excess Concentration"
        # block by header text, then iterate numbered tests below it.
        import re as _re
        block_start = None
        for r in range(1, sm.max_row + 1):
            for col in ("B", "L"):
                cv = sm[f"{col}{r}"].value
                if isinstance(cv, str) and cv.strip() == "Excess Concentration":
                    block_start = r + 1
                    break
            if block_start:
                break
        # Scan rows after the "Excess Concentration" header until we hit
        # the next labelled section ("Total Receivables Counted Towards
        # Borrowing Base"). Inside that range, any row whose B starts with
        # a number-and-period prefix is a concentration test.
        bridge_covenants = []
        if block_start:
            terminator_row = b_label_to_row.get("Total Receivables Counted Towards Borrowing Base")
            end_row = terminator_row if terminator_row else sm.max_row + 1
            for r in range(block_start, end_row):
                label_v = sm[f"B{r}"].value
                if not isinstance(label_v, str):
                    continue
                label_v = label_v.replace("\xa0", " ").strip()
                if not _re.match(r"^\d+\.\s", label_v):
                    continue
                short = _re.sub(r"^\d+\.\s+", "", label_v)[:80].rstrip()
                pct = sm[f"I{r}"].value
                limit = sm[f"J{r}"].value
                excess = sm[f"L{r}"].value
                actual_usd = sm[f"H{r}"].value
                if "Reserved" in short or (
                    not isinstance(pct, (int, float)) and not isinstance(actual_usd, (int, float))
                ):
                    continue
                bridge_covenants.append({
                    "test": short,
                    "actual_pct": round(float(pct) * 100, 2) if isinstance(pct, (int, float)) else None,
                    "limit_pct": round(float(limit) * 100, 2) if isinstance(limit, (int, float)) else None,
                    "excess_usd": round(float(excess), 2) if isinstance(excess, (int, float)) else 0,
                    "actual_usd": round(float(actual_usd), 2) if isinstance(actual_usd, (int, float)) else None,
                })
        out["covenants"] = bridge_covenants

    # NOTE: eligibility_summary.OVERALL_ELIGIBLE.eligible_balance_usd is
    # ~$160M but includes SOFOM-pledged collateral, so it's NOT the actual
    # Bridge-eligible balance. The real Bridge-eligible is computed below
    # from tape_end as bridge_collateral (eligible loans, net of
    # sofom_balance_usd). We still capture the gross overall_eligible for
    # diagnostic purposes.
    if "eligibility_summary" in wb.sheetnames:
        es = wb["eligibility_summary"]
        for row in es.iter_rows(values_only=True):
            if row and row[0] == "OVERALL_ELIGIBLE":
                out["overall_eligible_gross"] = float(row[4] or 0)
                # Total Receivables (gross, before eligibility filter) net of
                # SOFOM pledge would be the right "total receivables" for
                # Bridge. We derive it from the tape_end walk below; for now
                # capture the unscoped total from eligibility_summary for
                # comparison only.
                out["overall_total_receivables_gross"] = float(row[5] or 0)
                break

    if "historical_draws" in wb.sheetnames and "total_drawn" not in out:
        hd = wb["historical_draws"]
        total = 0.0
        for row in hd.iter_rows(min_row=2, values_only=True):
            if len(row) >= 3 and isinstance(row[2], (int, float)):
                total += row[2]
        out["total_drawn"] = total

    # Walk tape_end once to derive everything we need.
    # Per loan, capture company_id, credit_limit, industry, etc. so we can
    # compute concentration deductions on the second pass.
    if "tape_end" in wb.sheetnames:
        te = wb["tape_end"]
        hdr_row = next(te.iter_rows(values_only=True))
        cols = {}
        for name in ("balance_usd", "sofom_balance_usd", "elig", "days_past_due",
                     "company_id", "credit_limit_usd", "uw_score",
                     "naics_industry_id", "is_startup", "onboarding_date",
                     "country_code", "card_balance_usd", "jp_balance_usd",
                     "product", "name"):
            try:
                cols[name] = hdr_row.index(name)
            except ValueError:
                cols[name] = None

        # The Bridge tape_end contains all active loans, including Mexican
        # ones that are mostly pledged to SOFOM. The Bridge "portfolio" is
        # only the Bridge-owned portion per loan (balance_usd − sofom_balance_usd).
        loans: list[dict] = []          # elig=1 loans only — feeds BB calc
        bridge_total = 0.0              # all active loans, net of SOFOM pledge
        bridge_dq30 = 0.0
        bridge_accounts = 0
        bridge_total_recv_pre_elig = 0.0  # gross Bridge-owned, regardless of elig (= total receivables)
        # Stratifications — balance and account counts by various dimensions
        by_country: dict[str, dict] = {}   # ISO numeric code → {balance, accounts}
        by_dpd_bucket: dict[str, dict] = {} # 0 / 1-30 / 31-60 / 61-90 / 90+
        by_product: dict[str, float] = {"card": 0.0, "jeeves_pay": 0.0}
        top_debtors: dict = {}              # company_id → {balance, name}
        for row in te.iter_rows(min_row=2, values_only=True):
            if not row or cols["balance_usd"] is None:
                continue
            bal = row[cols["balance_usd"]]
            if not isinstance(bal, (int, float)) or bal <= 0:
                continue
            sof = row[cols["sofom_balance_usd"]] if cols["sofom_balance_usd"] is not None and isinstance(row[cols["sofom_balance_usd"]], (int, float)) else 0
            bridge_part = bal - sof
            if bridge_part <= 0:
                continue  # loan is fully in SOFOM, not Bridge
            bridge_accounts += 1
            bridge_total += bridge_part
            bridge_total_recv_pre_elig += bridge_part
            dpd = row[cols["days_past_due"]] if cols["days_past_due"] is not None and isinstance(row[cols["days_past_due"]], (int, float)) else 0
            if dpd >= 30:
                bridge_dq30 += bridge_part

            # Stratifications use the Bridge portion only
            cc = row[cols["country_code"]] if cols["country_code"] is not None else None
            cc_key = _country_label(cc)
            entry = by_country.setdefault(cc_key, {"balance": 0.0, "accounts": 0})
            entry["balance"] += bridge_part
            entry["accounts"] += 1

            bucket = _dpd_bucket(dpd)
            b_entry = by_dpd_bucket.setdefault(bucket, {"balance": 0.0, "accounts": 0})
            b_entry["balance"] += bridge_part
            b_entry["accounts"] += 1

            card_b = row[cols["card_balance_usd"]] if cols["card_balance_usd"] is not None and isinstance(row[cols["card_balance_usd"]], (int, float)) else 0
            jp_b = row[cols["jp_balance_usd"]] if cols["jp_balance_usd"] is not None and isinstance(row[cols["jp_balance_usd"]], (int, float)) else 0
            # The card/jp split is on TOTAL balance, but we only count the
            # Bridge portion — pro-rate using ratio of bridge_part / bal.
            ratio = bridge_part / bal if bal else 0
            by_product["card"] += card_b * ratio
            by_product["jeeves_pay"] += jp_b * ratio

            cid = row[cols["company_id"]] if cols["company_id"] is not None else None
            if cid is not None:
                name = row[cols["name"]] if cols["name"] is not None else None
                td = top_debtors.setdefault(cid, {"balance": 0.0, "name": name})
                td["balance"] += bridge_part

            if row[cols["elig"]] != 1:
                continue
            loans.append({
                "company_id": cid,
                "bridge_bal": bridge_part,
                "credit_limit": row[cols["credit_limit_usd"]] if cols["credit_limit_usd"] is not None else None,
                "uw_score": row[cols["uw_score"]] if cols["uw_score"] is not None else None,
                "industry": row[cols["naics_industry_id"]] if cols["naics_industry_id"] is not None else None,
                "startup": row[cols["is_startup"]] if cols["is_startup"] is not None else None,
                "dpd": dpd,
            })
        # Override the previously-misnamed values
        total_balance = bridge_total
        dq30_balance = bridge_dq30
        account_count = bridge_accounts

        # Persist stratifications
        out["by_country"] = {k: {"balance": round(v["balance"], 2), "accounts": v["accounts"]} for k, v in by_country.items()}
        out["by_dpd_bucket"] = {k: {"balance": round(v["balance"], 2), "accounts": v["accounts"]} for k, v in by_dpd_bucket.items()}
        out["by_product"] = {k: round(v, 2) for k, v in by_product.items()}
        # Top 10 debtors by balance
        top10 = sorted(top_debtors.items(), key=lambda x: -x[1]["balance"])[:10]
        out["top_debtors"] = [{"company_id": str(cid), "name": d["name"], "balance": round(d["balance"], 2)} for cid, d in top10]

    # Originations + roll-rate matrix from rollforward sheet
    if "rollforward" in wb.sheetnames:
        rf = wb["rollforward"]
        rf_hdr = [c.value for c in rf[1]]
        def _idx(name):
            try: return rf_hdr.index(name)
            except ValueError: return None
        card_col = _idx("card_disbursement_amount_usd")
        jp_col = _idx("jeeves_pay_disbursement_amount_usd")
        bop_bal_col = _idx("bop_balance_usd")
        eop_bal_col = _idx("eop_balance_usd")
        bop_dpd_col = _idx("bop_days_past_due")
        eop_dpd_col = _idx("eop_days_past_due")

        card_orig = 0.0
        jp_orig = 0.0
        # Roll-rate: counts and balances by (from_bucket, to_bucket)
        roll_count: dict = {}
        roll_balance: dict = {}
        for row in rf.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            if card_col is not None and isinstance(row[card_col], (int, float)):
                card_orig += row[card_col]
            if jp_col is not None and isinstance(row[jp_col], (int, float)):
                jp_orig += row[jp_col]
            # Only loans active at BOP feed roll rates
            bop_b = row[bop_bal_col] if bop_bal_col is not None and isinstance(row[bop_bal_col], (int, float)) else 0
            if bop_b <= 0:
                continue
            fb = _dpd_bucket(row[bop_dpd_col]) if bop_dpd_col is not None else "Current"
            tb = _dpd_bucket(row[eop_dpd_col]) if eop_dpd_col is not None else "Current"
            roll_count[(fb, tb)] = roll_count.get((fb, tb), 0) + 1
            roll_balance[(fb, tb)] = roll_balance.get((fb, tb), 0.0) + bop_b

        out["originations_period"] = {
            "card": round(card_orig, 2),
            "jeeves_pay": round(jp_orig, 2),
            "total": round(card_orig + jp_orig, 2),
        }
        # Row-normalized percentage: of loans in from_bucket at BOP, what %
        # ended up in each to_bucket at EOP.
        from_totals: dict[str, int] = {}
        for (fb, _), c in roll_count.items():
            from_totals[fb] = from_totals.get(fb, 0) + c
        out["roll_rate"] = {
            f"{fb}|{tb}": {
                "count": c,
                "pct": round(c / from_totals[fb] * 100, 2) if from_totals.get(fb) else 0.0,
                "balance_usd": round(roll_balance.get((fb, tb), 0.0), 2),
            }
            for (fb, tb), c in roll_count.items()
        }

        # BOP / EOP dates from the tape_start / tape_end sheets
        for sheet_name, key in (("tape_start", "bop_dt"), ("tape_end", "eop_dt")):
            if sheet_name not in wb.sheetnames:
                continue
            sh = wb[sheet_name]
            sh_hdr = [c.value for c in sh[1]]
            if "dt" not in sh_hdr:
                continue
            dt_col_i = sh_hdr.index("dt")
            # Take the dt from the first data row (all rows share the same dt
            # within a snapshot).
            for row in sh.iter_rows(min_row=2, values_only=True):
                if row and row[dt_col_i] is not None:
                    out[key] = str(row[dt_col_i])[:10]
                    break

    # Bank accounts (USD-converted)
    if "bank_accts" in wb.sheetnames:
        ba = wb["bank_accts"]
        bank_balances: dict[str, float] = {}  # country → USD
        for row in ba.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            country, _acct, currency, bal = row[0], row[1], row[2], row[3]
            if not isinstance(bal, (int, float)) or bal <= 0:
                continue
            # Approximate FX — for a glance-dashboard only. Cash is small
            # relative to receivables (<$3M typically for Bridge).
            if currency in ("USD", 840):
                bal_usd = float(bal)
            elif currency == "MXN":
                bal_usd = float(bal) / 17.35
            elif currency == "COP":
                bal_usd = float(bal) / 4100.0
            elif currency in ("BRL", 6540):
                bal_usd = float(bal) / 5.5
            else:
                bal_usd = float(bal)
            label = str(country) if country else "Unknown"
            bank_balances[label] = bank_balances.get(label, 0.0) + bal_usd
        out["bank_balances_by_country"] = {k: round(v, 2) for k, v in bank_balances.items()}

        gross_eligible = sum(l["bridge_bal"] for l in loans)
        out["bridge_collateral_gross"] = round(gross_eligible, 2)
        # Total receivables for Bridge = all Bridge-owned balance (net of
        # SOFOM pledge), regardless of eligibility. Use this as
        # total_receivables only if Summary didn't already provide one.
        if "total_receivables" not in out:
            out["total_receivables"] = round(bridge_total_recv_pre_elig, 2)
        # Don't overwrite Summary's canonical eligible if it was set.
        if "eligible" not in out:
            out["eligible"] = round(gross_eligible, 2)

        # ── Concentration breaches (informational) ────────────────────
        # These compute the *raw breach* per CIM Bridge BB tests using
        # tape_end columns. They are NOT subtracted from BB cap because
        # validation against the canonical BB file shows the BB pipeline
        # treats some of these as informational rather than deductive
        # (e.g., the UW-D test produces a $11M raw breach today but the
        # actual file's L60 doesn't reflect that reduction — availability
        # comes out at $3.5M, matching the facility cap, not collateral×0.85
        # minus deductions). When build_us.py emits a values JSON sidecar
        # we'll switch to the canonical L60 and stop guessing here.
        breaches: dict[str, float] = {}
        single_debtor_limit = min(4_500_000.0, gross_eligible * 0.05)
        by_debtor: dict = {}
        for l in loans:
            if l["company_id"] is not None:
                by_debtor[l["company_id"]] = by_debtor.get(l["company_id"], 0) + l["bridge_bal"]
        breaches["single_debtor"] = sum(max(0, b - single_debtor_limit) for b in by_debtor.values())
        top3 = sorted(by_debtor.values(), reverse=True)[:3]
        breaches["top3_debtors"] = max(0, sum(top3) - gross_eligible * 0.15)
        big_limit = sum(l["bridge_bal"] for l in loans if isinstance(l["credit_limit"], (int, float)) and l["credit_limit"] > 2_000_000)
        breaches["credit_limit_gt_2m"] = max(0, big_limit - gross_eligible * 0.30)
        score_d = sum(l["bridge_bal"] for l in loans if str(l["uw_score"]).upper() == "D")
        breaches["uw_score_d"] = max(0, score_d - gross_eligible * 0.10)
        by_industry: dict = {}
        for l in loans:
            if l["industry"] is not None:
                by_industry[l["industry"]] = by_industry.get(l["industry"], 0) + l["bridge_bal"]
        breaches["single_industry"] = sum(max(0, b - gross_eligible * 0.30) for b in by_industry.values())
        startup_bal = sum(l["bridge_bal"] for l in loans if l["startup"] in (1, True, "True", "true"))
        breaches["startups"] = max(0, startup_bal - gross_eligible * 0.20)

        # Only set eligible from tape if Summary's L60 didn't already (i.e.,
        # the canonical BB pipeline output is unavailable).
        out.setdefault("eligible", round(gross_eligible, 2))
        out["concentration_breaches"] = {k: round(v, 2) for k, v in breaches.items()}
        out["concentration_breaches_total"] = round(sum(breaches.values()), 2)

        # Bridge facility size: $75M (CIM contractual ceiling).
        # If Summary's canonical BB was read above, use that. Otherwise
        # fall back to gross_eligible × 0.85.
        BRIDGE_FACILITY_SIZE = 75_000_000.0
        out.setdefault("facility_size", BRIDGE_FACILITY_SIZE)
        if "borrowing_base" not in out:
            out["borrowing_base"] = round(gross_eligible * 0.85, 2)
            out["borrowing_base_estimate"] = True
        # Availability = min(BB, facility) − drawn. BB is the constraint
        # when BB < facility (overcollateralized); facility binds otherwise.
        bb = out.get("borrowing_base") or 0
        facility = out.get("facility_size") or 0
        drawn = out.get("total_drawn") or 0
        binding_cap = min(bb, facility) if (bb and facility) else (bb or facility)
        out["binding_cap"] = round(binding_cap, 2)
        out["availability"] = round(binding_cap - drawn, 2)

        out["portfolio_total"] = round(total_balance, 2)
        out["portfolio_accounts"] = account_count
        out["portfolio_dq30_pct"] = round(dq30_balance / total_balance * 100, 2) if total_balance else 0.0
        out["portfolio_as_of"] = "from tape_end"

    return out


def _read_mx_sofom(wb) -> dict:
    """Extract MX SOFOM BB metrics from the canonical Exhibit A sheet.

    Exhibit A is the BBVA SOFOM Advance Request. The BB pipeline writes
    cached values to specific cells which we read directly. The raw `tape`
    sheet is the broader Mexico loan pool ($213M total); Exhibit A's
    "Transferred Receivables" is the subset actually in the SOFOM facility
    (1.17B MXN ≈ $67M USD).

    Key Exhibit A cells:
      L6  Spot Exchange Rate USDMXN
      L31 Total Receivables (MXN)
      L33 Ineligible Receivables (MXN)
      L35 Total Eligible Receivables (MXN)
      L52 Total Receivables Counted Towards BB (MXN, post-concentration)
      L54 Unrestricted cash in Collection Account / Prerecycling (MXN)
      L58 Total Collateral Counted Towards BB (MXN)
      L65 Receivable Advance Rate
      L66 Cash Advance Rate
      L71 Facility Size (USD)
      L73 Receivable Borrowing Base (USD)
      L74 Cash Borrowing Base (USD)
      L75 Swap Contract Value Borrowing Base (USD)
    """
    out: dict = {
        "collection_cash_mxn": None,
        "total_drawn": None,
        "eligible": None,
        "borrowing_base": None,
        "availability": None,
        "facility_size": None,
        "portfolio_total": None,
        "portfolio_accounts": None,
        "portfolio_dq30_pct": None,
    }

    # ─ Exhibit A: BB calculation with cached values ─────────────────
    if "Exhibit A" in wb.sheetnames:
        ex = wb["Exhibit A"]

        def _cell(coord):
            v = ex[coord].value
            return float(v) if isinstance(v, (int, float)) else None

        usdmxn = _cell("L6") or 17.37
        out["usdmxn_rate"] = round(usdmxn, 4)

        total_receivables_mxn = _cell("L31")
        ineligible_mxn = _cell("L33")
        eligible_mxn = _cell("L35")
        counted_mxn = _cell("L52")
        collection_cash_mxn = _cell("L54")
        receivable_bb_usd = _cell("L73")
        cash_bb_usd = _cell("L74")
        swap_bb_usd = _cell("L75")
        facility_size_usd = _cell("L71")

        if total_receivables_mxn:
            out["total_receivables_mxn"] = round(total_receivables_mxn, 2)
            out["total_receivables"] = round(total_receivables_mxn / usdmxn, 2)
        if ineligible_mxn is not None:
            out["ineligible_mxn"] = round(ineligible_mxn, 2)
        if eligible_mxn is not None:
            out["eligible_mxn"] = round(eligible_mxn, 2)
            out["eligible"] = round(eligible_mxn / usdmxn, 2)
        if counted_mxn is not None:
            out["receivables_counted_mxn"] = round(counted_mxn, 2)
        if collection_cash_mxn is not None:
            out["collection_cash_mxn"] = round(collection_cash_mxn, 2)

        # Capture component cells for transparency.
        out["receivable_bb_usd"] = round(receivable_bb_usd or 0, 2)
        out["cash_bb_usd"] = round(cash_bb_usd or 0, 2)
        out["swap_bb_usd"] = round(swap_bb_usd or 0, 2)

        # Canonical "Borrowing Base" lives at L81 and "BB Excess / Deficit" at L87.
        # L81 applies the full BB pipeline (advance rates, NPY reserve at L63,
        # and the cash-offset Brian flagged). L73+L74+L75 is an intermediate
        # sum that overstates the BB by ~$6M.
        canonical_bb = _cell("L81")
        canonical_excess_deficit = _cell("L87")
        canonical_drawn = _cell("L83")
        canonical_unused = _cell("L89")
        if canonical_bb is not None:
            out["borrowing_base"] = round(canonical_bb, 2)
        if canonical_excess_deficit is not None:
            out["bb_excess_deficit"] = round(canonical_excess_deficit, 2)
        if canonical_drawn is not None and not os.environ.get("SOFOM_DRAWN_OVERRIDE"):
            out["total_drawn"] = round(canonical_drawn, 2)
        if canonical_unused is not None:
            out["unused_facility"] = round(canonical_unused, 2)
        if facility_size_usd:
            out["facility_size"] = facility_size_usd

        # SOFOM concentration covenants (Exhibit A rows 40-50)
        # H = Actual ($ MXN), I = Percentage, J = Limit (decimal), K = Limit $, L = Excess ($)
        _SOFOM_TEST_LABELS = {
            40: "Credit limit > $1.5M",
            41: "Single Obligor (<$40M pool)",
            42: "Single Obligor (otherwise)",
            43: "Top 3 Obligors (<$40M pool)",
            44: "Top 3 Obligors (otherwise)",
            45: "Risk Score D",
            46: "Start-Ups",
            47: "Single Industry",
            48: "High-Risk Industry",
            49: "High-Risk Provinces",
            50: "Onboarded last 6 months",
        }
        covenants = []
        for r, label in _SOFOM_TEST_LABELS.items():
            actual = ex[f"H{r}"].value
            pct = ex[f"I{r}"].value
            limit = ex[f"J{r}"].value
            excess = ex[f"L{r}"].value
            if not isinstance(pct, (int, float)) and not isinstance(actual, (int, float)):
                continue
            covenants.append({
                "test": label,
                "actual_pct": round(float(pct) * 100, 2) if isinstance(pct, (int, float)) else None,
                "limit_pct": round(float(limit) * 100, 2) if isinstance(limit, (int, float)) else None,
                "excess_mxn": round(float(excess), 2) if isinstance(excess, (int, float)) else 0,
                "actual_mxn": round(float(actual), 2) if isinstance(actual, (int, float)) else None,
            })
        out["covenants"] = covenants

    # ─ historical_draws: USD-denominated ─────────────────────────────
    # File may lag real-time (today's draw may not be in latest published
    # BB). SOFOM_DRAWN_OVERRIDE env var lets ops nudge until regenerated.
    if "historical_draws" in wb.sheetnames:
        hd = wb["historical_draws"]
        total = 0.0
        for row in hd.iter_rows(min_row=2, values_only=True):
            if len(row) >= 3 and isinstance(row[2], (int, float)):
                total += row[2]
        override = os.environ.get("SOFOM_DRAWN_OVERRIDE")
        if override:
            try:
                total = float(override)
                out["total_drawn_override_active"] = True
            except ValueError:
                pass
        out["total_drawn"] = round(total, 2)

    # Portfolio total = Exhibit A's "Total Receivables" (Transferred
    # Receivables, USD). The raw tape sums to $213M but Exhibit A's
    # subset is what's actually in the SOFOM facility ($67M USD).
    if out.get("total_receivables") is not None:
        out["portfolio_total"] = out["total_receivables"]

    # Tape walk for stratifications + DQ30 + accounts. The SOFOM tape is
    # multi-snapshot (3 different `dt` values, ~620-645 rows each) — we
    # filter to the latest dt only to avoid 3x-counting balances.
    if "tape" in wb.sheetnames:
        tape = wb["tape"]
        hdr = [c.value for c in tape[1]]
        tcols = {}
        for name in ("sofom_balance_usd", "balance_usd", "days_past_due",
                     "country_code", "card_balance_usd", "jp_balance_usd",
                     "company_id", "name", "dt"):
            try:
                tcols[name] = hdr.index(name)
            except ValueError:
                tcols[name] = None

        # First pass: find max dt
        max_dt = None
        dt_col = tcols.get("dt")
        if dt_col is not None:
            for row in tape.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                d = row[dt_col]
                if d is not None and (max_dt is None or d > max_dt):
                    max_dt = d
            if max_dt is not None:
                out["tape_as_of"] = str(max_dt)

        sof_col = tcols["sofom_balance_usd"]
        dpd_col = tcols["days_past_due"]
        if sof_col is not None and dpd_col is not None:
            tape_total = 0.0
            tape_dq30 = 0.0
            tape_accts = 0
            by_country: dict[str, dict] = {}
            by_dpd_bucket: dict[str, dict] = {}
            by_product: dict[str, float] = {"card": 0.0, "jeeves_pay": 0.0}
            top_debtors: dict = {}
            for row in tape.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                # Only the latest snapshot — earlier dts are historical
                if dt_col is not None and max_dt is not None and row[dt_col] != max_dt:
                    continue
                sof = row[sof_col]
                if not isinstance(sof, (int, float)) or sof <= 0:
                    continue
                bal = row[tcols["balance_usd"]] if tcols["balance_usd"] is not None and isinstance(row[tcols["balance_usd"]], (int, float)) else sof
                tape_total += sof
                tape_accts += 1
                dpd = row[dpd_col] if isinstance(row[dpd_col], (int, float)) else 0
                if dpd >= 30:
                    tape_dq30 += sof

                cc = row[tcols["country_code"]] if tcols["country_code"] is not None else None
                cc_key = _country_label(cc)
                c_entry = by_country.setdefault(cc_key, {"balance": 0.0, "accounts": 0})
                c_entry["balance"] += sof
                c_entry["accounts"] += 1

                bucket = _dpd_bucket(dpd)
                b_entry = by_dpd_bucket.setdefault(bucket, {"balance": 0.0, "accounts": 0})
                b_entry["balance"] += sof
                b_entry["accounts"] += 1

                card_b = row[tcols["card_balance_usd"]] if tcols["card_balance_usd"] is not None and isinstance(row[tcols["card_balance_usd"]], (int, float)) else 0
                jp_b = row[tcols["jp_balance_usd"]] if tcols["jp_balance_usd"] is not None and isinstance(row[tcols["jp_balance_usd"]], (int, float)) else 0
                ratio = sof / bal if bal else 0
                by_product["card"] += card_b * ratio
                by_product["jeeves_pay"] += jp_b * ratio

                cid = row[tcols["company_id"]] if tcols["company_id"] is not None else None
                if cid is not None:
                    name_v = row[tcols["name"]] if tcols["name"] is not None else None
                    td = top_debtors.setdefault(cid, {"balance": 0.0, "name": name_v})
                    td["balance"] += sof

            if tape_total > 0:
                out["portfolio_dq30_pct"] = round(tape_dq30 / tape_total * 100, 2)
                ratio = (out.get("portfolio_total") or tape_total) / tape_total
                out["portfolio_accounts"] = int(round(tape_accts * ratio))
                out["by_country"] = {k: {"balance": round(v["balance"], 2), "accounts": v["accounts"]} for k, v in by_country.items()}
                out["by_dpd_bucket"] = {k: {"balance": round(v["balance"], 2), "accounts": v["accounts"]} for k, v in by_dpd_bucket.items()}
                out["by_product"] = {k: round(v, 2) for k, v in by_product.items()}
                top10 = sorted(top_debtors.items(), key=lambda x: -x[1]["balance"])[:10]
                out["top_debtors"] = [{"company_id": str(cid), "name": d["name"], "balance": round(d["balance"], 2)} for cid, d in top10]

    # Bank accounts (USD-converted) for SOFOM
    if "bank_accts" in wb.sheetnames:
        ba = wb["bank_accts"]
        bank_balances: dict[str, float] = {}
        usdmxn = out.get("usdmxn_rate") or 17.35
        for row in ba.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            country, _acct, currency, bal = row[0], row[1], row[2], row[3]
            if not isinstance(bal, (int, float)) or bal <= 0:
                continue
            if currency in ("USD", 840):
                bal_usd = float(bal)
            elif currency == "MXN":
                bal_usd = float(bal) / usdmxn
            elif currency == "COP":
                bal_usd = float(bal) / 4100.0
            elif currency in ("BRL", 6540):
                bal_usd = float(bal) / 5.5
            else:
                bal_usd = float(bal)
            label = str(country) if country else "Unknown"
            bank_balances[label] = bank_balances.get(label, 0.0) + bal_usd
        out["bank_balances_by_country"] = {k: round(v, 2) for k, v in bank_balances.items()}

    # Availability = min(BB, facility_size) − drawn. The constraint that
    # binds is whichever is smaller — that's the whole point of having a BB.
    bb = out.get("borrowing_base")
    facility = out.get("facility_size")
    drawn = out.get("total_drawn")
    if bb is not None and drawn is not None:
        binding_cap = min(bb, facility) if facility is not None else bb
        out["binding_cap"] = round(binding_cap, 2)
        out["availability"] = round(binding_cap - drawn, 2)

    return out


def fetch_bb_metrics() -> dict:
    """Pull latest US Bridge + MX SOFOM BB from Drive."""
    svc = dfu.get_drive_service()

    bridge_file = dfu.search_latest_bb(svc, "Bridge Borrowing Base")
    sofom_file = dfu.search_latest_bb(svc, "SOFOM Borrowing Base")

    us: dict = {}
    if bridge_file:
        log.info("US Bridge BB: %s", bridge_file["name"])
        wb = _download_xlsx(svc, bridge_file["id"])
        us = _read_us_bridge(wb)
        us["source_file"] = bridge_file["name"]
        us["source_mtime"] = bridge_file.get("modifiedTime")
    else:
        log.warning("No US Bridge BB file found on Drive")

    mx: dict = {}
    if sofom_file:
        log.info("MX SOFOM BB: %s", sofom_file["name"])
        wb = _download_xlsx(svc, sofom_file["id"])
        mx = _read_mx_sofom(wb)
        # Fall back to CSV-text-based collection_cash_mxn — the SOFOM
        # Prerecycling line is reliably parseable via regex even though the
        # rest of the CSV parser is fragile.
        if mx.get("collection_cash_mxn") is None:
            text = dfu.fetch_bb_text(sofom_file["id"])
            parsed = dfu.parse_sofom_bb(text)
            cc = parsed.get("collection_cash_mxn", 0)
            mx["collection_cash_mxn"] = float(cc) if cc else None
        mx["source_file"] = sofom_file["name"]
        mx["source_mtime"] = sofom_file.get("modifiedTime")
    else:
        log.warning("No MX SOFOM BB file found on Drive")

    return {"us_bridge": us, "mx_sofom": mx}


def fetch_roll_rate_from_redshift() -> dict:
    """Compute month-over-month DPD roll rates from loc_tape.

    Picks the two most recent month-end snapshots and computes per-loan
    DPD transitions. Loans active (balance > 0) at the prior month-end
    are the cohort; their DPD at the latest month-end determines the
    destination bucket. Loans absent or with zero balance at EOP go to
    "Paid/Charged-off".
    """
    try:
        from redshift_util import connect
    except Exception as e:
        log.warning("redshift_util import failed: %s", e)
        return {}

    try:
        conn = connect()
        cur = conn.cursor()
        # Redshift doesn't allow month/year intervals in many contexts.
        # Use DATEADD + LAST_DAY-equivalent date math.
        cur.execute("""
            SELECT DISTINCT dt
            FROM capital_markets_dm.loc_tape
            WHERE dt = DATEADD('day', -1, DATEADD('month', 1, date_trunc('month', dt)))::date
              AND dt < date_trunc('month', current_date)::date
            ORDER BY dt DESC
            LIMIT 2
        """)
        rows = cur.fetchall()
        if len(rows) < 2:
            log.warning("Need 2 month-end snapshots for roll rate; got %d", len(rows))
            conn.close()
            return {}
        eop_dt, bop_dt = str(rows[0][0]), str(rows[1][0])
        log.info("Roll rate period: BOP=%s -> EOP=%s", bop_dt, eop_dt)

        # Bucket labels are zero-prefixed so they sort naturally in alpha order.
        # The dashboard strips the prefix via findReplaceByRegex.
        # Cohort filter: active at BOP AND not already charged off.
        # Destination separates Paid Off (balance went to 0 normally) from
        # Charged Off (charge_off_flag/charge_off_dt set).
        cur.execute(f"""
            WITH bop AS (
                SELECT loan_id, balance_usd, days_past_due
                FROM capital_markets_dm.loc_tape
                WHERE dt = '{bop_dt}'
                  AND balance_usd > 0
                  AND (charge_off_flag IS NULL OR charge_off_flag = 0)
            ),
            eop AS (
                SELECT loan_id, balance_usd, days_past_due,
                       charge_off_flag, charge_off_dt
                FROM capital_markets_dm.loc_tape
                WHERE dt = '{eop_dt}'
            )
            SELECT
                CASE
                    WHEN b.days_past_due IS NULL OR b.days_past_due <= 0 THEN '01 Current'
                    WHEN b.days_past_due <= 30 THEN '02 1-30 DPD'
                    WHEN b.days_past_due <= 60 THEN '03 31-60 DPD'
                    WHEN b.days_past_due <= 90 THEN '04 61-90 DPD'
                    ELSE '05 90+ DPD'
                END AS from_bucket,
                CASE
                    WHEN e.charge_off_flag = 1 OR e.charge_off_dt IS NOT NULL THEN '07 Charged Off'
                    WHEN e.loan_id IS NULL OR e.balance_usd IS NULL OR e.balance_usd <= 0 THEN '06 Paid Off'
                    WHEN e.days_past_due IS NULL OR e.days_past_due <= 0 THEN '01 Current'
                    WHEN e.days_past_due <= 30 THEN '02 1-30 DPD'
                    WHEN e.days_past_due <= 60 THEN '03 31-60 DPD'
                    WHEN e.days_past_due <= 90 THEN '04 61-90 DPD'
                    ELSE '05 90+ DPD'
                END AS to_bucket,
                COUNT(*) AS cnt,
                SUM(b.balance_usd) AS bop_balance_usd
            FROM bop b
            LEFT JOIN eop e ON b.loan_id = e.loan_id
            GROUP BY 1, 2
        """)
        roll: dict = {}
        from_totals: dict[str, int] = {}
        for fb, tb, cnt, bal in cur.fetchall():
            roll[(fb, tb)] = {"count": int(cnt), "bop_balance_usd": float(bal or 0)}
            from_totals[fb] = from_totals.get(fb, 0) + int(cnt)
        conn.close()
    except Exception as e:
        log.warning("Redshift roll-rate query failed: %s", e)
        return {}

    if not roll:
        return {}

    out_roll = {
        f"{fb}|{tb}": {
            "count": stats["count"],
            "pct": round(stats["count"] / from_totals[fb] * 100, 2) if from_totals.get(fb) else 0.0,
            "bop_balance_usd": round(stats["bop_balance_usd"], 2),
        }
        for (fb, tb), stats in roll.items()
    }
    return {"roll_rate": out_roll, "bop_dt": bop_dt, "eop_dt": eop_dt}


# Portfolio metrics are now computed inside _read_us_bridge from the BB
# file's tape_end sheet — same tape the BB engine uses, charge-offs already
# excluded (balance_usd > 0). No Redshift dependency for the Grafana view.


def _combine_stratifications(us: dict, mx: dict) -> dict:
    """Merge Bridge + SOFOM stratifications into one global view.

    Sums balance + accounts across both facilities for each dimension.
    Top debtors are combined, then re-ranked across the union.
    """
    g: dict = {
        "by_country": {},
        "by_dpd_bucket": {},
        "by_product": {},
        "bank_balances_by_country": {},
    }
    for block in (us, mx):
        for country, stats in (block.get("by_country") or {}).items():
            if not isinstance(stats, dict):
                continue
            entry = g["by_country"].setdefault(country, {"balance": 0.0, "accounts": 0})
            entry["balance"] += stats.get("balance") or 0
            entry["accounts"] += stats.get("accounts") or 0
        for bucket, stats in (block.get("by_dpd_bucket") or {}).items():
            if not isinstance(stats, dict):
                continue
            entry = g["by_dpd_bucket"].setdefault(bucket, {"balance": 0.0, "accounts": 0})
            entry["balance"] += stats.get("balance") or 0
            entry["accounts"] += stats.get("accounts") or 0
        for product, bal in (block.get("by_product") or {}).items():
            g["by_product"][product] = g["by_product"].get(product, 0.0) + (bal or 0)
        for country, bal in (block.get("bank_balances_by_country") or {}).items():
            g["bank_balances_by_country"][country] = g["bank_balances_by_country"].get(country, 0.0) + (bal or 0)

    # Combined top debtors — sum balances by company_id across facilities,
    # re-rank, take top 10.
    debtor_pool: dict = {}
    for block in (us, mx):
        for td in (block.get("top_debtors") or []):
            if not isinstance(td, dict):
                continue
            cid = td.get("company_id") or ""
            entry = debtor_pool.setdefault(cid, {"name": td.get("name"), "balance": 0.0})
            entry["balance"] += td.get("balance") or 0
            if not entry["name"]:
                entry["name"] = td.get("name")
    top10 = sorted(debtor_pool.items(), key=lambda x: -x[1]["balance"])[:10]
    g["top_debtors"] = [{"company_id": cid, "name": d["name"], "balance": round(d["balance"], 2)} for cid, d in top10]

    # Period originations (Bridge has these from rollforward; SOFOM build
    # doesn't currently expose them — sum what we have).
    card = (us.get("originations_period") or {}).get("card") or 0
    jp = (us.get("originations_period") or {}).get("jeeves_pay") or 0
    mx_card = (mx.get("originations_period") or {}).get("card") or 0
    mx_jp = (mx.get("originations_period") or {}).get("jeeves_pay") or 0
    g["originations_period"] = {
        "card": round(card + mx_card, 2),
        "jeeves_pay": round(jp + mx_jp, 2),
        "total": round(card + jp + mx_card + mx_jp, 2),
    }

    # Roll-rate: Bridge rollforward (covers all geos including Mexico). SOFOM
    # has no rollforward sheet so this is the only source.
    g["roll_rate"] = us.get("roll_rate") or {}
    g["bop_dt"] = us.get("bop_dt")
    g["eop_dt"] = us.get("eop_dt")

    # Round balances after summation.
    for d in (g["by_country"], g["by_dpd_bucket"]):
        for k, v in d.items():
            v["balance"] = round(v["balance"], 2)
    g["by_product"] = {k: round(v, 2) for k, v in g["by_product"].items()}
    g["bank_balances_by_country"] = {k: round(v, 2) for k, v in g["bank_balances_by_country"].items()}
    return g


def main() -> int:
    bb = fetch_bb_metrics()
    us = bb.get("us_bridge", {}) or {}
    mx = bb.get("mx_sofom", {}) or {}

    # Strip per-facility strat blocks once we've combined them — Brian
    # wants strats global only, not facility-scoped.
    global_strats = _combine_stratifications(us, mx)

    # Override BB-rollforward roll rate (4-day Bridge-only window) with
    # the Redshift month-over-month roll rate.
    rs_roll = fetch_roll_rate_from_redshift()
    if rs_roll.get("roll_rate"):
        global_strats["roll_rate"] = rs_roll["roll_rate"]
        global_strats["bop_dt"] = rs_roll.get("bop_dt")
        global_strats["eop_dt"] = rs_roll.get("eop_dt")
    for block in (us, mx):
        for k in ("by_country", "by_dpd_bucket", "by_product",
                  "top_debtors", "bank_balances_by_country",
                  "originations_period", "roll_rate", "bop_dt", "eop_dt"):
            block.pop(k, None)

    state = {
        "us_bridge": us,
        "mx_sofom":  mx,
        "portfolio": global_strats,
        "updated_at": datetime.now().isoformat(),
    }
    OUT_FILE.write_text(json.dumps(state, indent=2), encoding="utf-8")
    log.info("Wrote %s", OUT_FILE)

    # Merge into _cap_markets_state.json so the gateway scrape picks it up.
    writer = BACKEND_DIR / "scripts" / "cap_markets_metrics_writer.py"
    result = subprocess.run(
        [sys.executable, str(writer)],
        capture_output=True, text=True, timeout=30,
    )
    if result.returncode != 0:
        log.error("cap_markets_metrics_writer failed: %s", result.stderr.strip()[-500:])
        return 1
    log.info("Cap markets metrics merged into _cap_markets_state.json")
    return 0


def _state_age_seconds(path, ts_field):
    """Seconds since path's ts_field timestamp, or None if missing/unreadable."""
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return (datetime.now() - datetime.fromisoformat(data[ts_field])).total_seconds()
    except Exception:
        return None


def run_loop():
    """Cron-supervisor entry point — refresh metrics every hour forever."""
    import time
    interval = int(os.environ.get("CAP_MARKETS_REFRESH_SECONDS", "3600"))
    log.info("cap-markets-refresh cron starting (interval=%ds)", interval)
    while True:
        # Skip-if-fresh: every gateway restart re-instantiates this cron, which
        # would otherwise immediately re-run ~100s of Redshift queries. During
        # that burst the gateway's async /health probe gets starved → the
        # supervisor false-kills the gateway → restart → fresh burst (the
        # 2026-06-16 flap). Reusing still-fresh state lets a restarted gateway
        # pass its health checks before doing heavy work.
        age = _state_age_seconds(OUT_FILE, "updated_at")
        if age is not None and age < interval:
            wait = interval - age
            log.info("BB metrics fresh (%.0fs old < %ds) — skipping run, sleeping %.0fs", age, interval, wait)
            time.sleep(wait)
            continue
        try:
            main()
        except Exception as e:
            log.exception("cap-markets-refresh iteration failed: %s", e)
        time.sleep(interval)


if __name__ == "__main__":
    sys.exit(main())
