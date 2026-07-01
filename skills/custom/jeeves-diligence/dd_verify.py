#!/usr/bin/env python3
"""dd_verify.py — pre-delivery verification gate for diligence work.

Run this BEFORE reporting any tracker update, VDR upload, or DD package as
done. "The API call returned" is not verification — this script checks the
result as the counterparty will see it.

Usage:
    python dd_verify.py --tracker <path/to/tracker.xlsx>   # verify every link in a tracker
    python dd_verify.py --folder <DRIVE_FOLDER_ID>         # verify every file in a DD/VDR folder
    python dd_verify.py --tracker <xlsx> --folder <ID>     # both

Checks:
    --tracker: extracts every hyperlink (cell hyperlinks + URLs in cell text)
               and verifies each one resolves: Drive links via the API
               (exists, not trashed), other URLs via HTTP.
    --folder:  lists the folder and verifies every file actually opens —
               downloads xlsx/docx/pdf and parses them (catches corrupted
               uploads), checks native Google files for non-trashed status,
               and flags duplicate copies of the same document (same name
               modulo date stamps) and empty files.

Exit code 0 = all checks passed. Non-zero = failures listed in output.

Requires env vars: GOOGLE_CLIENT_ID, GOOGLE_CLIENT_SECRET, GOOGLE_REFRESH_TOKEN
(loaded automatically from backend/.env when run outside the sandbox).
"""

import argparse
import io
import os
import re
import sys
import zipfile

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_SCRIPT_DIR, '..', '_shared'))
sys.path.insert(0, os.path.join(_SCRIPT_DIR, '..', 'google-drive'))

from env_loader import load_env
load_env()
from google_auth import get_credentials
from upload_to_drive import normalize_doc_name

_DRIVE_ID_RE = re.compile(
    r'(?:docs|drive)\.google\.com/(?:file/d/|document/d/|spreadsheets/d/|presentation/d/|drive/folders/)([A-Za-z0-9_-]{10,})')
_URL_RE = re.compile(r'https?://[^\s<>"\')\]]+')

PASS, FAIL, WARN = 'PASS', 'FAIL', 'WARN'


def _drive_service():
    from googleapiclient.discovery import build
    return build('drive', 'v3', credentials=get_credentials())


def _check_drive_id(service, file_id):
    """Return (status, detail) for a Drive file ID."""
    try:
        meta = service.files().get(
            fileId=file_id, fields='id,name,trashed,mimeType,size').execute()
    except Exception as e:
        if '404' in str(e):
            return FAIL, 'Drive file not found (404) — broken or deleted link'
        return FAIL, f'Drive error: {str(e)[:120]}'
    if meta.get('trashed'):
        return FAIL, f"Drive file is in the trash: {meta.get('name')}"
    return PASS, meta.get('name', '')


def _check_http(url):
    try:
        import requests
        r = requests.head(url, allow_redirects=True, timeout=15)
        if r.status_code == 405:  # some servers reject HEAD
            r = requests.get(url, allow_redirects=True, timeout=15, stream=True)
        if r.status_code < 400:
            return PASS, f'HTTP {r.status_code}'
        return FAIL, f'HTTP {r.status_code}'
    except Exception as e:
        return FAIL, f'request failed: {str(e)[:100]}'


def _extract_tracker_links(tracker_path):
    """Yield (location, url) for every hyperlink and in-text URL in the workbook."""
    import openpyxl
    wb = openpyxl.load_workbook(tracker_path)
    seen = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                loc = f'{ws.title}!{cell.coordinate}'
                if cell.hyperlink and cell.hyperlink.target:
                    url = cell.hyperlink.target
                    if url not in seen:
                        seen.add(url)
                        yield loc, url
                if isinstance(cell.value, str):
                    for url in _URL_RE.findall(cell.value):
                        if url not in seen:
                            seen.add(url)
                            yield loc, url


def verify_tracker(tracker_path, service):
    results = []
    links = list(_extract_tracker_links(tracker_path))
    if not links:
        results.append((WARN, 'tracker', tracker_path, 'no hyperlinks found in workbook — is that expected?'))
        return results
    for loc, url in links:
        m = _DRIVE_ID_RE.search(url)
        if m:
            status, detail = _check_drive_id(service, m.group(1))
        elif url.startswith('http'):
            status, detail = _check_http(url)
        else:
            status, detail = WARN, 'not a URL I can verify'
        results.append((status, loc, url[:80], detail))
    return results


def _verify_file_content(service, meta):
    """Download an office/pdf file and confirm it parses. Returns (status, detail)."""
    from googleapiclient.http import MediaIoBaseDownload
    name = meta['name']
    size = int(meta.get('size') or 0)
    if size == 0:
        return FAIL, 'file is 0 bytes'

    buf = io.BytesIO()
    try:
        dl = MediaIoBaseDownload(buf, service.files().get_media(fileId=meta['id']))
        done = False
        while not done:
            _, done = dl.next_chunk()
    except Exception as e:
        return FAIL, f'download failed: {str(e)[:100]}'
    buf.seek(0)
    data = buf.getvalue()

    lower = name.lower()
    try:
        if lower.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return PASS, f'opens OK ({len(sheets)} sheets: {", ".join(sheets[:5])})'
        if lower.endswith(('.docx', '.pptx')):
            zf = zipfile.ZipFile(io.BytesIO(data))
            bad = zf.testzip()
            return (FAIL, f'corrupt archive member: {bad}') if bad else (PASS, 'opens OK')
        if lower.endswith('.pdf'):
            if data[:5] == b'%PDF-':
                return PASS, f'valid PDF ({size / 1024:.0f} KB)'
            return FAIL, 'does not start with %PDF — corrupt'
        if lower.endswith(('.csv', '.txt', '.md')):
            data.decode('utf-8-sig', errors='strict')
            return PASS, f'decodes OK ({size / 1024:.0f} KB)'
        return PASS, f'{size / 1024:.0f} KB (content not parsed for this type)'
    except Exception as e:
        return FAIL, f'file does not open: {str(e)[:120]}'


def verify_folder(folder_id, service):
    results = []
    try:
        res = service.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields='files(id,name,mimeType,size,modifiedTime)',
            pageSize=200,
        ).execute()
    except Exception as e:
        return [(FAIL, 'folder', folder_id, f'cannot list folder: {str(e)[:120]}')]

    files = res.get('files', [])
    if not files:
        results.append((WARN, 'folder', folder_id, 'folder is empty'))
        return results

    # Duplicate detection: same document (name modulo dates) more than once
    by_key = {}
    for f in files:
        if f['mimeType'] == 'application/vnd.google-apps.folder':
            continue
        by_key.setdefault(normalize_doc_name(f['name']), []).append(f)
    for key, group in by_key.items():
        if len(group) > 1:
            names = '; '.join(g['name'] for g in group)
            results.append((FAIL, 'duplicates', key, f'{len(group)} copies of the same document: {names}'))

    for f in files:
        if f['mimeType'] == 'application/vnd.google-apps.folder':
            results.append((WARN, 'subfolder', f['name'], 'subfolder present — flatten unless intentional'))
            continue
        if f['mimeType'].startswith('application/vnd.google-apps.'):
            # Native Google file — existence + non-trashed already confirmed by listing
            results.append((PASS, 'file', f['name'], 'native Google file, listed OK'))
            continue
        status, detail = _verify_file_content(service, f)
        results.append((status, 'file', f['name'], detail))
    return results


def main():
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')

    parser = argparse.ArgumentParser(description='Verify diligence tracker links and DD folder contents')
    parser.add_argument('--tracker', help='Path to a tracker .xlsx to verify all links in')
    parser.add_argument('--folder', help='Drive folder ID to verify all files in')
    args = parser.parse_args()

    if not args.tracker and not args.folder:
        parser.error('provide --tracker and/or --folder')

    service = _drive_service()
    results = []

    if args.tracker:
        if not os.path.isfile(args.tracker):
            print(f'ERROR: tracker not found: {args.tracker}', file=sys.stderr)
            sys.exit(1)
        print(f'\n=== Verifying tracker links: {os.path.basename(args.tracker)} ===')
        results += verify_tracker(args.tracker, service)

    if args.folder:
        print(f'\n=== Verifying folder contents: {args.folder} ===')
        results += verify_folder(args.folder, service)

    fails = [r for r in results if r[0] == FAIL]
    warns = [r for r in results if r[0] == WARN]

    print()
    for status, kind, subject, detail in results:
        print(f'[{status}] {kind}: {subject}')
        print(f'       {detail}')

    print(f'\n{"=" * 60}')
    print(f'TOTAL: {len(results)} checks — {len(results) - len(fails) - len(warns)} pass, {len(warns)} warn, {len(fails)} FAIL')
    if fails:
        print('\nVERIFICATION FAILED — fix the items above before reporting this work as done.')
        sys.exit(1)
    print('All checks passed.')


if __name__ == '__main__':
    main()
