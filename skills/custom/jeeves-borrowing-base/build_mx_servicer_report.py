#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Jeeves Mexico SPV Monthly Servicer Report - April 2026
Period: BOP=2026-04-01, EOP=2026-04-30
"""

import os, sys, datetime as dt, importlib.util, warnings
from pathlib import Path
warnings.filterwarnings('ignore')

OUTPUTS_PATH = os.environ.get('OUTPUTS_PATH', '/mnt/user-data/outputs')
BB_DIR = '/mnt/skills/custom/jeeves-borrowing-base'

# Load BB modules via importlib (Windows-safe path handling)
def load_module(name, fpath):
    spec = importlib.util.spec_from_file_location(name, fpath)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

ru = load_module('redshift_util', BB_DIR + '/redshift_util.py')
ru.ensure_deps('openpyxl')

import pandas as pd
from openpyxl import load_workbook
import datetime as dt

elig_mod = load_module('eligibility', BB_DIR + '/eligibility.py')
calc_elig = elig_mod.calculate_eligibility_fields_sofom

# Config
BOP_DATE    = dt.date(2026, 4, 1)
EOP_DATE    = dt.date(2026, 4, 30)
REPORT_DATE = EOP_DATE

# SQL
TAPE_SQL = Path(BB_DIR + '/sql/data_tape_sofom.sql').read_text(encoding='utf-8')
RF_SQL   = Path('/mnt/skills/custom/cfo-org-kb/sql/loc_acct_rollforward_sofom.sql').read_text(encoding='utf-8')

# Connect
print("Connecting to Redshift...")
con = ru.connect()

# 1. SOFOM tapes
print(f"Querying tape BOP={BOP_DATE}...")
df_bop = pd.read_sql_query(TAPE_SQL.format(BOP_DATE.isoformat()), con)
df_bop = calc_elig(df_bop)
print(f"  {len(df_bop)} rows")

print(f"Querying tape EOP={EOP_DATE}...")
df_eop = pd.read_sql_query(TAPE_SQL.format(EOP_DATE.isoformat()), con)
df_eop = calc_elig(df_eop)
print(f"  {len(df_eop)} rows")

# 2. Rollforward
print("Querying rollforward...")
df_rf = pd.read_sql_query(RF_SQL.format(BOP_DATE.isoformat(), EOP_DATE.isoformat()), con)
print(f"  {len(df_rf)} rows")
rf_agg = df_rf.sum(numeric_only=True)

# 3. Revenue
print("Querying revenue...")
rev_sql = """
select
    sum(lt.fee_amount) as total_interest_fees_mxn,
    sum(lt.late_payment_penalty_amount) as total_late_fees_mxn,
    sum(lt.jeeves_pay_fee_amount) as total_jp_fees_mxn
from capital_markets_dm.loc_tape lt
inner join dms_mysql_jeeves_raw.company_settings cs
    on cs.companyId = lt.company_id
    and cs.settingKey = 'SOFOM_JPMORGAN_ENABLED'
    and cs.settingValue = 'on'
where lt.dt between '{}'::date and '{}'::date
  and lt.is_in_repayment is False
  and lt.charge_off_flag is False
""".format(BOP_DATE.isoformat(), EOP_DATE.isoformat())
df_rev = pd.read_sql_query(rev_sql, con)

# 4. DQ buckets at EOP
print("Querying DQ buckets...")
dq_sql = """
select
    case
        when days_past_due = 0 then 'Current'
        when days_past_due between 1 and 30 then '1-30 DPD'
        when days_past_due between 31 and 60 then '31-60 DPD'
        when days_past_due between 61 and 90 then '61-90 DPD'
        when days_past_due between 91 and 120 then '91-120 DPD'
        else '120+'
    end as dq_bucket,
    count(*) as cnt,
    sum(greatest(0, lt2.sofom_balance)) as balance_mxn
from (
    select
        lt3.company_id, lt3.days_past_due,
        greatest(0, lt3.balance - coalesce(ba.assignment_balance, 0)) as sofom_balance
    from capital_markets_dm.loc_tape lt3
    inner join dms_mysql_jeeves_raw.company_settings cs2
        on cs2.companyId = lt3.company_id
        and cs2.settingKey = 'SOFOM_JPMORGAN_ENABLED'
        and cs2.settingValue = 'on'
    left join (
        select l4.company_id, l4.loan_id, l4.balance as assignment_balance
        from capital_markets_dm.loc_tape l4
        inner join dms_mysql_jeeves_raw.company_settings cs4
            on cs4.companyId = l4.company_id
            and cs4.settingKey = 'SOFOM_JPMORGAN_ENABLED'
        where l4.dt = date(cs4.updatedat)
    ) ba on ba.company_id = lt3.company_id and ba.loan_id = lt3.loan_id
    where lt3.dt = '{}'
      and lt3.is_in_repayment is False
      and lt3.charge_off_flag is False
) lt2
group by 1
""".format(EOP_DATE.isoformat())
df_dq = pd.read_sql_query(dq_sql, con)
print(f"  DQ buckets found: {df_dq['dq_bucket'].tolist()}")

con.close()
print("All Redshift queries complete.\n")

# Compute metrics
total_recv_mxn   = float(df_eop['sofom_balance'].sum())
total_recv_count = len(df_eop)
eligible_eop     = df_eop[df_eop['elig'] == 1]
eligible_mxn     = float(eligible_eop['sofom_balance'].sum())
ineligible_mxn   = total_recv_mxn - eligible_mxn

bop_mxn  = float(df_bop['sofom_balance'].sum())
bop_neg  = float(df_bop[df_bop['sofom_balance'] < 0]['sofom_balance'].sum())
bop_adj  = bop_mxn + abs(bop_neg)
eop_neg  = float(df_eop[df_eop['sofom_balance'] < 0]['sofom_balance'].sum())
eop_adj  = total_recv_mxn + abs(eop_neg)

disbursements_mxn  = float(rf_agg.get('card_disbursement_amount', 0) + rf_agg.get('jeeves_pay_disbursement_amount', 0))
payments_mxn       = float(rf_agg.get('payment_amount', 0))
cashback_mxn       = float(rf_agg.get('cashback_amount', 0))
late_fees_mxn      = float(rf_agg.get('late_payment_penalty_amount', 0))
jp_interest_mxn    = float(rf_agg.get('jeeves_pay_fee_amount', 0))
loan_alloc_mxn     = float(rf_agg.get('loan_allocation_amount', 0))
adjustments_mxn    = float(rf_agg.get('adjustment_amount', 0) + rf_agg.get('fx_adjustment_amount', 0))
charge_offs_mxn    = float(rf_agg.get('charge_off', 0))
billing_ccy_mxn    = float(rf_agg.get('usd_delta_currency_switch', 0))
modified_recv_mxn  = float(rf_agg.get('repayment', 0))

# Concentrations
top_obligors     = df_eop.groupby('company_id')['sofom_balance'].sum().sort_values(ascending=False)
df_over_limit    = df_eop[df_eop['credit_limit_usd'] > 1_500_000]
conc_a_actual    = float(df_over_limit['sofom_balance'].sum())
conc_a_pct       = conc_a_actual / total_recv_mxn if total_recv_mxn else 0
conc_a_limit_mxn = eligible_mxn * 0.30
conc_a_excess    = max(0.0, conc_a_actual - conc_a_limit_mxn)

conc_b_actual    = float(top_obligors.iloc[0]) if len(top_obligors) else 0
conc_b_pct       = conc_b_actual / eligible_mxn if eligible_mxn else 0
conc_b_limit_mxn = eligible_mxn * 0.05
conc_b_excess    = max(0.0, conc_b_actual - conc_b_limit_mxn)

conc_c_actual    = float(top_obligors.iloc[:3].sum()) if len(top_obligors) >= 3 else float(top_obligors.sum())
conc_c_pct       = conc_c_actual / eligible_mxn if eligible_mxn else 0
conc_c_limit_mxn = eligible_mxn * 0.10
conc_c_excess    = max(0.0, conc_c_actual - conc_c_limit_mxn)

df_d_rated       = df_eop[df_eop['uw_score'] == 'D']
conc_d_actual    = float(df_d_rated['sofom_balance'].sum())
conc_d_pct       = conc_d_actual / eligible_mxn if eligible_mxn else 0
conc_d_limit_mxn = eligible_mxn * 0.05
conc_d_excess    = max(0.0, conc_d_actual - conc_d_limit_mxn)

df_startup       = df_eop[df_eop['is_startup'] == 1]
conc_e_actual    = float(df_startup['sofom_balance'].sum())
conc_e_pct       = conc_e_actual / eligible_mxn if eligible_mxn else 0
conc_e_limit_mxn = eligible_mxn * 0.20
conc_e_excess    = max(0.0, conc_e_actual - conc_e_limit_mxn)

avg_daily_bal_mxn    = (bop_adj + eop_adj) / 2
int_collections_mxn  = float(df_rev['total_interest_fees_mxn'].iloc[0] or 0)
fee_collections_mxn  = float((df_rev['total_late_fees_mxn'].iloc[0] or 0) + (df_rev['total_jp_fees_mxn'].iloc[0] or 0))
total_nonprincipal_mxn = int_collections_mxn + fee_collections_mxn

dq_lookup   = df_dq.set_index('dq_bucket')['balance_mxn'].to_dict()
dq_current  = float(dq_lookup.get('Current', 0))
dq_1_30     = float(dq_lookup.get('1-30 DPD', 0))
dq_31_60    = float(dq_lookup.get('31-60 DPD', 0))
dq_61_90    = float(dq_lookup.get('61-90 DPD', 0))
dq_91_120   = float(dq_lookup.get('91-120 DPD', 0))
dq_120plus  = float(dq_lookup.get('120+', 0))
dq_total    = total_recv_mxn

lista_cols  = ['company_id','name','sofom_balance','fee_amount','onboarding_date']
df_lista    = df_eop[lista_cols].copy()
df_lista.columns = ['numero de contrato','Nombre del Deudor JVS Mexico',
                    'new spend','interest','Fecha de firma']

# Build workbook
template_path   = '/mnt/user-data/workspace/servicer_template.xlsx'
output_filename = f'Jeeves Mexico Servicer Report - {REPORT_DATE.strftime("%m-%d-%Y")}.xlsx'
output_path     = os.path.join(OUTPUTS_PATH, output_filename)

print("Loading template...")
wb = load_workbook(template_path)

def fset(ws, label, col_offset, value, max_rows=None):
    """Find cell with label and set value at col_offset to right."""
    for row in ws.iter_rows(max_row=max_rows or ws.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and label.lower() in cell.value.lower():
                ws.cell(cell.row, cell.column + col_offset).value = value
                return True
    return False

# --- Exhibit A ---
ws_a = wb['Exhibit A']
fset(ws_a, 'Begin Date', 10, dt.datetime(BOP_DATE.year, BOP_DATE.month, BOP_DATE.day))
fset(ws_a, 'End Date',   10, dt.datetime(EOP_DATE.year, EOP_DATE.month, EOP_DATE.day))
fset(ws_a, 'Spot Exchange Rate', 10, '[FILL IN]')

rf_items = [
    ('Begin Portolio Balance', bop_mxn),
    ('Beginning Balance',      bop_adj),
    ('(+) Disbursements',      disbursements_mxn),
    ('(+) Payments',           payments_mxn),
    ('(+) Cashback',           cashback_mxn),
    ('(+) Late Fees',          late_fees_mxn),
    ('(+) Jeeves Pay Interest',jp_interest_mxn),
    ('(+) Payment Reallocation', loan_alloc_mxn),
    ('(+) Adjustments',        adjustments_mxn),
    ('(+) Charge Offs',        charge_offs_mxn),
    ('(+) Billing Currency',   billing_ccy_mxn),
    ('(+) Modified',           modified_recv_mxn),
    ('End Portolio Balance',   total_recv_mxn),
    ('Ending Balance',         eop_adj),
]
for label, val in rf_items:
    fset(ws_a, label, 10, float(val))

neg_hits = 0
for row in ws_a.iter_rows(max_row=30):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and '(-) Negative Balances' in cell.value:
            ws_a.cell(cell.row, cell.column + 10).value = float(abs(bop_neg) if neg_hits == 0 else abs(eop_neg))
            neg_hits += 1

fset(ws_a, 'Total Receivables', 9, total_recv_count)
fset(ws_a, 'Total Receivables', 10, total_recv_mxn)
fset(ws_a, 'Ineligible Receivables', 10, ineligible_mxn)
fset(ws_a, 'Total Eligible', 10, eligible_mxn)

for row in ws_a.iter_rows():
    for cell in row:
        v = cell.value
        if not (v and isinstance(v, str)):
            continue
        c, r = cell.column, cell.row
        if '(a)' in v and 'credit limit' in v.lower():
            for off, val in enumerate([conc_a_actual, conc_a_pct, 0.30, conc_a_limit_mxn, conc_a_excess], 6):
                ws_a.cell(r, c+off).value = float(val)
        elif '(b)' in v and '7.50%' in v:
            for off, val in enumerate([0.0, 0.0, 0.075, eligible_mxn*0.075, 0.0], 6):
                ws_a.cell(r, c+off).value = float(val)
        elif '(b)' in v and '5.00%' in v:
            for off, val in enumerate([conc_b_actual, conc_b_pct, 0.05, conc_b_limit_mxn, conc_b_excess], 6):
                ws_a.cell(r, c+off).value = float(val)
        elif '(c)' in v and '15%' in v:
            for off, val in enumerate([0.0, 0.0, 0.15, eligible_mxn*0.15, 0.0], 6):
                ws_a.cell(r, c+off).value = float(val)
        elif '(c)' in v and '10.00%' in v:
            for off, val in enumerate([conc_c_actual, conc_c_pct, 0.10, conc_c_limit_mxn, conc_c_excess], 6):
                ws_a.cell(r, c+off).value = float(val)
        elif '(d)' in v and 'Score of D' in v:
            for off, val in enumerate([conc_d_actual, conc_d_pct, 0.05, conc_d_limit_mxn, conc_d_excess], 6):
                ws_a.cell(r, c+off).value = float(val)
        elif '(e)' in v and 'Start-Up' in v:
            for off, val in enumerate([conc_e_actual, conc_e_pct, 0.20, conc_e_limit_mxn, conc_e_excess], 6):
                ws_a.cell(r, c+off).value = float(val)

# --- Exhibit B ---
ws_b = wb['Exhibit B']
fset(ws_b, 'Begin Date', 2, dt.datetime(BOP_DATE.year, BOP_DATE.month, BOP_DATE.day))
fset(ws_b, 'End Date',   2, dt.datetime(EOP_DATE.year, EOP_DATE.month, EOP_DATE.day))
fset(ws_b, 'Spot Exchange Rate', 2, '[FILL IN]')
fset(ws_b, 'Average daily balance', 5, avg_daily_bal_mxn)
fset(ws_b, 'Borrower Interest Collections', 5, int_collections_mxn)
fset(ws_b, 'Borrower Other Fee Collections', 5, fee_collections_mxn)
fset(ws_b, 'Total non-principal', 5, total_nonprincipal_mxn)

dq_map = {
    'Current':    (dq_current, dq_current/dq_total if dq_total else 0),
    '1-30 DPD':   (dq_1_30,   dq_1_30/dq_total if dq_total else 0),
    '31-60 DPD':  (dq_31_60,  dq_31_60/dq_total if dq_total else 0),
    '61-90 DPD':  (dq_61_90,  dq_61_90/dq_total if dq_total else 0),
    '91-120 DPD': (dq_91_120, dq_91_120/dq_total if dq_total else 0),
    '120+':       (dq_120plus,dq_120plus/dq_total if dq_total else 0),
}
for row in ws_b.iter_rows():
    for cell in row:
        v = str(cell.value).strip() if cell.value else ''
        if v in dq_map:
            bal, pct = dq_map[v]
            ws_b.cell(cell.row, cell.column + 5).value = float(bal)
            ws_b.cell(cell.row, cell.column + 6).value = float(pct)
        elif v == 'Total' and cell.row > 10:
            ws_b.cell(cell.row, cell.column + 5).value = float(dq_total)
            ws_b.cell(cell.row, cell.column + 6).value = 1.0

fset(ws_b, 'Available Amounts for distribution', 5, '[FILL IN - Monex/CxC collection account]')

# --- tape tab: BOP + EOP ---
ws_tape = wb['tape']
for row in ws_tape.iter_rows(min_row=2, max_row=ws_tape.max_row):
    for cell in row:
        cell.value = None

df_tape = pd.concat([df_bop, df_eop], ignore_index=True)
for col_idx, col_name in enumerate(df_tape.columns, 1):
    ws_tape.cell(1, col_idx, col_name)

print(f"Writing tape tab ({len(df_tape)} rows)...")
for row_idx in range(len(df_tape)):
    for col_idx, val in enumerate(df_tape.iloc[row_idx], 1):
        if hasattr(val, 'item'):
            val = val.item()
        if val is not None and isinstance(val, float) and val != val:
            val = None
        ws_tape.cell(row_idx + 2, col_idx, val)

# --- lista tab ---
ws_lista = wb['lista']
for row in ws_lista.iter_rows(min_row=2, max_row=ws_lista.max_row):
    for cell in row:
        cell.value = None
for col_idx, col_name in enumerate(df_lista.columns, 1):
    ws_lista.cell(1, col_idx, col_name)
for row_idx in range(len(df_lista)):
    for col_idx, val in enumerate(df_lista.iloc[row_idx], 1):
        if hasattr(val, 'item'):
            val = val.item()
        if val is not None and isinstance(val, float) and val != val:
            val = None
        ws_lista.cell(row_idx + 2, col_idx, val)

# --- bank_accts tab ---
ws_bank = wb['bank_accts']
for row in ws_bank.iter_rows(min_row=2):
    for cell in row:
        if cell.column == 3:
            cell.value = '[FILL IN]'

# --- Save ---
os.makedirs(OUTPUTS_PATH, exist_ok=True)
print(f"Saving...")
wb.save(output_path)
print(f"Done: {output_path}")

# Summary
pf = lambda e: 'PASS' if e == 0 else '!! BREACH !!'
print(f"""
=== Jeeves Mexico Servicer Report - April 2026 ===

EXHIBIT A
  BOP Balance:        {bop_adj:>18,.0f} MXN
  EOP Total Recv:     {eop_adj:>18,.0f} MXN  ({total_recv_count} accounts)
  Eligible:           {eligible_mxn:>18,.0f} MXN
  Ineligible:         {ineligible_mxn:>18,.0f} MXN

ROLLFORWARD (Apr 1 - Apr 30)
  Disbursements:      {disbursements_mxn:>18,.0f} MXN
  Payments:           {payments_mxn:>18,.0f} MXN
  Cashback:           {cashback_mxn:>18,.0f} MXN
  Late Fees:          {late_fees_mxn:>18,.0f} MXN
  Charge-Offs:        {charge_offs_mxn:>18,.0f} MXN

CONCENTRATION TESTS (vs Eligible Balance)
  (a) >$1.5M CL:    {conc_a_pct*100:5.1f}% vs 30.0%  [{pf(conc_a_excess)}]
  (b) Top obligor:  {conc_b_pct*100:5.1f}% vs  5.0%  [{pf(conc_b_excess)}]
  (c) Top 3:        {conc_c_pct*100:5.1f}% vs 10.0%  [{pf(conc_c_excess)}]
  (d) D-rated:      {conc_d_pct*100:5.1f}% vs  5.0%  [{pf(conc_d_excess)}]
  (e) Startups:     {conc_e_pct*100:5.1f}% vs 20.0%  [{pf(conc_e_excess)}]

DQ AT EOP (Apr 30)
  Current:     {dq_current:>16,.0f} MXN  ({dq_current/dq_total*100:.1f}%)
  1-30 DPD:    {dq_1_30:>16,.0f} MXN  ({dq_1_30/dq_total*100:.2f}%)
  31-60 DPD:   {dq_31_60:>16,.0f} MXN  ({dq_31_60/dq_total*100:.2f}%)
  61-90 DPD:   {dq_61_90:>16,.0f} MXN  ({dq_61_90/dq_total*100:.2f}%)
  91-120 DPD:  {dq_91_120:>16,.0f} MXN  ({dq_91_120/dq_total*100:.2f}%)
  120+:        {dq_120plus:>16,.0f} MXN  ({dq_120plus/dq_total*100:.2f}%)

[FILL IN BEFORE SENDING]
  1. FX spot rate (USDMXN Apr 30) - Exhibit A & B, Spot Exchange Rate cells
  2. JPM + Monex bank balances    - bank_accts tab, bal column
  3. FX hedge MTM values          - hedge tab
  4. April draws (if any)         - historical_draws tab
  5. Collection acct avail. amt   - Exhibit B, Available Amounts row
""")
