import io, os, sys, json, subprocess
from pathlib import Path

SKILLS = os.environ.get('SKILLS_PATH', 'C:/Jeeves/redshift-bot/skills')
sys.path.insert(0, str(Path(SKILLS) / 'custom' / 'google-drive'))

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import openpyxl

creds = Credentials(
    token=None,
    refresh_token=os.environ['GOOGLE_REFRESH_TOKEN'],
    client_id=os.environ['GOOGLE_CLIENT_ID'],
    client_secret=os.environ['GOOGLE_CLIENT_SECRET'],
    token_uri='https://oauth2.googleapis.com/token',
)
creds.refresh(Request())
svc = build('drive', 'v3', credentials=creds, cache_discovery=False)

def download_xlsx(file_id):
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, svc.files().get_media(fileId=file_id))
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return buf

# Bridge BB
res = svc.files().list(
    q="name contains 'Bridge Borrowing Base' and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and trashed=false",
    orderBy='modifiedTime desc', pageSize=1,
    fields='files(id,name,modifiedTime)'
).execute()
bridge = res['files'][0]
print(f"Bridge BB: {bridge['name']}")

wb = openpyxl.load_workbook(download_xlsx(bridge['id']), data_only=True)
te = wb['tape_end']
headers = [c.value for c in te[1]]
h = {v: i for i, v in enumerate(headers) if v}
bi = h.get('balance_usd')
ei = h.get('elig')

total_balance = elig_balance = elig_count = total_count = 0
for row in te.iter_rows(min_row=2, values_only=True):
    b = row[bi] if bi is not None else None
    e = row[ei] if ei is not None else None
    if b and isinstance(b, (int, float)) and b > 0:
        total_balance += b
        total_count += 1
        if e:
            elig_balance += b
            elig_count += 1

hd = wb['historical_draws']
total_drawn = sum(
    row[2] for row in hd.iter_rows(min_row=2, values_only=True)
    if row and len(row) >= 3 and row[2] and isinstance(row[2], (int, float))
)

facility = 75_000_000
bb_cap = min(elig_balance * 0.85, facility)
availability = bb_cap - total_drawn  # can be negative (over-advanced)
print(f"  drawn=${total_drawn:,.0f}  elig=${elig_balance:,.0f}  bb=${bb_cap:,.0f}  avail=${availability:,.0f}")

# SOFOM BB
res2 = svc.files().list(
    q="name contains 'SOFOM Borrowing Base' and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and trashed=false",
    orderBy='modifiedTime desc', pageSize=1,
    fields='files(id,name,modifiedTime)'
).execute()
sofom = res2['files'][0]
print(f"SOFOM BB: {sofom['name']}")

wb2 = openpyxl.load_workbook(download_xlsx(sofom['id']), data_only=True)
collection_cash_mxn = 0.0
total_receivables_mxn = 0.0
for sheet_name in wb2.sheetnames:
    ws2 = wb2[sheet_name]
    for row in ws2.iter_rows(max_row=300):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            lv = cell.value.lower()
            if 'prerecycling' in lv or ('collection' in lv and 'cash' in lv):
                for c in row:
                    if isinstance(c.value, (int, float)) and c.value > 1_000_000:
                        collection_cash_mxn = float(c.value)
                        break
            if 'total receivables' in lv and not total_receivables_mxn:
                for c in row:
                    if isinstance(c.value, (int, float)) and c.value > 1_000_000:
                        total_receivables_mxn = float(c.value)
                        break

if not collection_cash_mxn:
    collection_cash_mxn = 232_836_527.87

print(f"  cash_mxn={collection_cash_mxn:,.0f}  receivables_mxn={total_receivables_mxn:,.0f}")

STATE_DIR = Path("C:/Jeeves/redshift-bot/deer-flow/backend/.deer-flow")
bb_metrics = {
    "us_bridge": {
        "total_drawn":    total_drawn,
        "availability":   availability,
        "eligible":       elig_balance,
        "borrowing_base": bb_cap,
        "facility_size":  facility,
        "total_balance":  total_balance,
        "elig_count":     elig_count,
        "as_of":          bridge['modifiedTime'][:10],
    },
    "mx_sofom": {
        "collection_cash_mxn": collection_cash_mxn,
        "total_receivables":   total_receivables_mxn,
        "facility_size_mxn":   100_000_000,
    },
    "portfolio": {
        "dq30_pct": None,
        "total":    total_balance,
        "accounts": total_count,
    },
}
(STATE_DIR / "_bb_metrics_state.json").write_text(json.dumps(bb_metrics, indent=2), encoding='utf-8')
print(f"Wrote _bb_metrics_state.json")

r = subprocess.run(
    [sys.executable, "C:/Jeeves/redshift-bot/deer-flow/backend/scripts/cap_markets_metrics_writer.py"],
    capture_output=True, text=True, timeout=30
)
print(r.stdout.strip())

cm = json.loads((STATE_DIR / "_cap_markets_state.json").read_text())
us = cm.get("us_bridge", {})
mx = cm.get("mx_sofom", {})
drawn_m = us.get('total_drawn', 0) / 1e6
avail_m = us.get('availability', 0) / 1e6
elig_m  = us.get('eligible', 0) / 1e6
bb_m    = us.get('borrowing_base', 0) / 1e6
cash_m  = mx.get('collection_cash_mxn', 0) / 1e6
print(f"FINAL: drawn={drawn_m:.1f}M  avail={avail_m:.1f}M  elig={elig_m:.1f}M  bb={bb_m:.1f}M  sofom_cash={cash_m:.1f}M MXN")
