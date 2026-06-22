#!/usr/bin/env python3
"""Diligence tracker monitor cron.

Runs inside the gateway via cron_supervisor (expects a blocking run_loop()).

Purpose (Brian, 2026-06-15): keep the consolidated Diligence Master Tracker
ALWAYS in sync with its sources so nothing ever drifts. The master is a pure
derived roll-up — never hand-edited. This cron is the thing that regenerates it.

Each cycle (business hours, weekdays):
  1. Snapshot the file inventory of all four counterparty folders + the Kroll
     source request tracker's status cells.
  2. Diff against last snapshot (state JSON).
  3. If anything changed, run diligence_master_builder.py (live-mirrors the
     Kroll source tracker, rebuilds + re-uploads the master in place) and DM
     Brian a concise drift report.

State: backend/.deer-flow/_diligence_tracker_state.json
"""

import json
import logging
import os
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

_SCRIPT_DIR = Path(__file__).resolve().parent
_BACKEND_DIR = _SCRIPT_DIR.parent.parent.parent / "backend"
_STATE_FILE = _BACKEND_DIR / ".deer-flow" / "_diligence_tracker_state.json"
_BUILDER = _SCRIPT_DIR / "diligence_master_builder.py"
_SHARED = _SCRIPT_DIR.parent / "_shared"

CHECK_INTERVAL_SECS = 2 * 3600          # poll every 2h
BUSINESS_HOURS = range(7, 20)           # 07:00–19:00 local

# Workstream sources to monitor
FOLDERS = {
    "Vista VCP/PwC (VDR)": "1JiMVbmEKQxKN_e5TXY2MWY3LXJGqv48P",
    "Vista Legal (PSK)":   "1qK4gkzHHkC6rMJV7aUBTu-UHntbCAoWY",
    "NB / Akin uploads":   "1yae9sldBwAvj3G8fRYY2EdGtE33VPKfX",
    "Kroll":               "1amlOqm423CkTVYWqukx9HmXE5kTCZiZC",
}
KROLL_SOURCE_TRACKER = "19G5hv8M8SVZ9YwLeMPCRiQn669jxXfmW"


def _load_state() -> dict:
    try:
        with open(_STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_state(state: dict) -> None:
    _STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(_STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)


def _drive():
    sys.path.insert(0, str(_SHARED))
    from google_auth import get_credentials
    from googleapiclient.discovery import build
    return build("drive", "v3", credentials=get_credentials())


def _snapshot() -> dict:
    """Inventory each folder (id->name+modified) + Kroll source statuses."""
    svc = _drive()
    snap = {"folders": {}, "kroll_status": {}}
    for label, fid in FOLDERS.items():
        items = {}
        page = None
        while True:
            resp = svc.files().list(
                q=f"'{fid}' in parents and trashed=false",
                fields="nextPageToken, files(id,name,modifiedTime)",
                pageSize=200, pageToken=page,
                includeItemsFromAllDrives=True, supportsAllDrives=True,
            ).execute()
            for f in resp.get("files", []):
                items[f["id"]] = {"name": f["name"], "modified": f.get("modifiedTime", "")}
            page = resp.get("nextPageToken")
            if not page:
                break
        snap["folders"][label] = items

    # Kroll source tracker status cells
    try:
        import io, tempfile
        from googleapiclient.http import MediaIoBaseDownload
        from openpyxl import load_workbook
        tmp = os.path.join(tempfile.gettempdir(), "_kroll_src_cron.xlsx")
        req = svc.files().get_media(fileId=KROLL_SOURCE_TRACKER)
        fh = io.FileIO(tmp, "wb"); d = MediaIoBaseDownload(fh, req); done = False
        while not done:
            _, done = d.next_chunk()
        fh.close()
        ws = load_workbook(tmp)["Kroll Request Tracker"]
        for r in range(1, ws.max_row + 1):
            ref = ws.cell(r, 1).value
            if isinstance(ref, str) and (ref.startswith("P1-") or ref.startswith("P2-")):
                snap["kroll_status"][ref.strip()] = (ws.cell(r, 4).value or "").strip()
    except Exception as e:
        logger.warning("[diligence-tracker] kroll status snapshot failed: %s", e)
    return snap


def _diff(old: dict, new: dict) -> list:
    changes = []
    of = (old or {}).get("folders", {})
    for label, items in new.get("folders", {}).items():
        prev = of.get(label, {})
        for fid, meta in items.items():
            if fid not in prev:
                changes.append(f"[{label}] NEW file: {meta['name']}")
            elif meta["modified"] != prev[fid].get("modified"):
                changes.append(f"[{label}] updated: {meta['name']}")
        for fid, meta in prev.items():
            if fid not in items:
                changes.append(f"[{label}] removed: {meta['name']}")
    ok = (old or {}).get("kroll_status", {})
    for ref, st in new.get("kroll_status", {}).items():
        if ref in ok and ok[ref] != st:
            changes.append(f"[Kroll] {ref} status: {ok[ref]} -> {st}")
    return changes


def _dm_owner(text: str) -> None:
    token = os.environ.get("SLACK_BOT_TOKEN")
    owner = os.environ.get("SLACK_OWNER_USER_ID")
    if not token or not owner:
        return
    try:
        from slack_sdk import WebClient
        client = WebClient(token=token)
        ch = client.conversations_open(users=[owner])["channel"]["id"]
        client.chat_postMessage(channel=ch, text=text)
    except Exception as e:
        logger.warning("[diligence-tracker] Slack DM failed: %s", e)


def _rebuild_master() -> str:
    env = dict(os.environ, PYTHONPATH=str(_SHARED), PYTHONIOENCODING="utf-8")
    result = subprocess.run(
        [sys.executable, str(_BUILDER)],
        capture_output=True, text=True, timeout=900, env=env, cwd=str(_SCRIPT_DIR),
    )
    if result.returncode != 0:
        raise RuntimeError(f"builder exited {result.returncode}: {(result.stderr or result.stdout)[-600:]}")
    return result.stdout[-400:]


def _run_once() -> None:
    state = _load_state()
    new = _snapshot()
    changes = _diff(state, new)
    if changes or not state.get("last_built"):
        out = _rebuild_master()
        state["last_built"] = datetime.now().isoformat()
        if changes:
            body = "\n".join(f"• {c}" for c in changes[:40])
            _dm_owner(
                ":card_index_dividers: *Diligence tracker updated* — sources changed, master regenerated:\n"
                f"```{body}```"
            )
        logger.info("[diligence-tracker] rebuilt master (%d changes). %s", len(changes), out)
    state["last_check"] = datetime.now().isoformat()
    state["snapshot"] = new
    _save_state(state)


def run_loop() -> None:
    logger.info("[diligence-tracker] cron started (poll every %ds, business hours)", CHECK_INTERVAL_SECS)
    while True:
        try:
            now = datetime.now()
            if now.weekday() < 5 and now.hour in BUSINESS_HOURS:
                _run_once()
        except Exception:
            raise  # let cron_supervisor alert + restart with backoff
        time.sleep(CHECK_INTERVAL_SECS)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    if "--once" in sys.argv:
        _run_once()
    else:
        run_loop()
