#!/usr/bin/env python3
"""Build the monthly Portfolio Report.

Usage:
    python build_portfolio_report.py --date 2026-04-01 --template-id <DRIVE_FILE_ID>

--date is the REPORT DATE (1st of the month). The report covers the previous
month's data:
  --date 2026-04-01 → EOP=2026-03-31, BOP=2026-02-28
  File: Portfolio Report - 20260401.xlsx → uploaded to Portfolio Reporting/202604/

Steps:
  1. Run LOC tape + rollforward queries for EOP (day before report date)
  2. Run GWC mods query (repayment plans only: loan_reference_number LIKE 'RPP%')
  3. Download previous month's report from Drive as template
  4. Copy current period's formula values (col L) to prior period column (col N)
     in Summary tabs so MoM comparison works
  5. Replace data tabs: loc, rollforward, mods
  6. Save as Portfolio Report - {YYYYMM}01.xlsx

Output saved to $OUTPUTS_PATH/Portfolio Report - {YYYYMM}01.xlsx
"""

import argparse
import datetime as dt
import os
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def _ensure_deps():
    try:
        import psycopg2, pandas, openpyxl  # noqa: F401
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-q',
                               'psycopg2-binary', 'pandas', 'openpyxl'])


def _connect():
    import psycopg2
    return psycopg2.connect(
        host=os.environ['REDSHIFT_HOST'],
        port=int(os.environ['REDSHIFT_PORT']),
        dbname=os.environ['REDSHIFT_DB'],
        user=os.environ['REDSHIFT_USER'],
        password=os.environ['REDSHIFT_PASSWORD'],
        sslmode='require',
        sslrootcert='disable',
    )


def _download_template(file_id: str, dest_path: str) -> None:
    """Download a file from Google Drive."""
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
    import io

    creds = Credentials(
        token=None,
        refresh_token=os.environ['GOOGLE_REFRESH_TOKEN'],
        token_uri='https://oauth2.googleapis.com/token',
        client_id=os.environ['GOOGLE_CLIENT_ID'],
        client_secret=os.environ['GOOGLE_CLIENT_SECRET'],
    )
    service = build('drive', 'v3', credentials=creds)
    request = service.files().get_media(fileId=file_id)
    fh = io.FileIO(dest_path, 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
        if status:
            pct = int(status.progress() * 100)
            if pct % 25 == 0:
                print(f"    Download {pct}%")
    fh.close()


def _shift_summary_mom(ws) -> int:
    """Copy current-period values (col L) to prior-period column (col N).

    Walks every row in a Summary sheet. If col L has a formula, copies
    the *computed value* from data_only into col N (hard-coded number).
    Returns the number of cells shifted.
    """
    shifted = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=12, max_col=12):
        cell_l = row[0]
        if cell_l.value is not None:
            # Col N = column 14
            cell_n = ws.cell(row=cell_l.row, column=14)
            # We need the evaluated value, not the formula.
            # Since we load with data_only=False (to preserve formulas),
            # we read the cached value from a data_only copy.
            # The caller handles this by passing cached values.
            shifted += 1
    return shifted


def _shift_summary_mom_with_values(ws_formula, cached_values: dict) -> int:
    """Copy cached computed values from col L into col N (hard-coded).

    cached_values: {row_number: value} from a data_only=True load.
    """
    shifted = 0
    for row_num, value in cached_values.items():
        if value is not None:
            ws_formula.cell(row=row_num, column=14, value=value)
            shifted += 1
    return shifted


def _get_col_l_values(ws_data_only) -> dict:
    """Extract col L values from a data_only worksheet."""
    values = {}
    for row in ws_data_only.iter_rows(min_row=1, max_row=ws_data_only.max_row,
                                       min_col=12, max_col=12):
        cell = row[0]
        if cell.value is not None:
            values[cell.row] = cell.value
    return values


def _replace_data_sheet(wb, sheet_name: str, df, *, include_index: bool = False) -> None:
    """Replace a data sheet in the workbook with new DataFrame content.

    Clears the existing sheet and writes new data in place, preserving
    the sheet's position in the workbook tab order.

    Args:
        wb: openpyxl Workbook.
        sheet_name: Name of the sheet to replace.
        df: pandas DataFrame with the new data.
        include_index: If True, write the DataFrame index as column A
            (matches the pandas default ``to_excel`` behavior).  The
            ``loc`` and ``rollforward`` tabs in the real portfolio report
            have a pandas index column; ``mods`` does not.
    """
    from openpyxl.utils.dataframe import dataframe_to_rows

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Clear all existing data (rows and columns) without deleting the sheet
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    else:
        # Sheet doesn't exist — append at end
        ws = wb.create_sheet(sheet_name)

    for r_idx, row in enumerate(dataframe_to_rows(df, index=include_index, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)


def main():
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')

    parser = argparse.ArgumentParser(description='Build monthly Portfolio Report')
    parser.add_argument('--date', type=str, required=True,
                        help='End-of-month date YYYY-MM-DD (e.g. 2026-03-31)')
    parser.add_argument('--template-id', type=str, required=True,
                        help='Google Drive file ID of the previous month\'s Portfolio Report')
    args = parser.parse_args()

    _ensure_deps()
    import pandas as pd
    import openpyxl

    # Compute dates
    # --date is the report date (1st of month), e.g. 2026-04-01.
    # The report covers the PREVIOUS month:
    #   EOP = day before report date (2026-03-31) — last day of data
    #   BOP = last day of the month before that (2026-02-28)
    # File is named with the report date and uploaded to the report month folder.
    yesterday = dt.date.today() - dt.timedelta(days=1)
    report_date = dt.datetime.strptime(args.date, '%Y-%m-%d').date()
    date_end = report_date - dt.timedelta(days=1)  # EOP = day before report date

    if date_end >= dt.date.today():
        print(f"ERROR: EOP date {date_end} (derived from report date {report_date}) "
              f"is today or in the future. Redshift data is only available through "
              f"yesterday ({yesterday}).", file=sys.stderr)
        sys.exit(1)

    # BOP = last day of month before EOP's month
    date_beg = date_end.replace(day=1) - dt.timedelta(days=1)
    month_str = report_date.strftime('%Y%m')  # folder/filename uses report date month

    print(f"Portfolio Report: report_date={report_date}, BOP={date_beg}, EOP={date_end}")

    # ── Step 1: Query LOC tape + rollforward ───────────────────────
    tape_sql = open(os.path.join(SCRIPT_DIR, 'sql', 'data_tape.sql')).read()
    rf_sql = open(os.path.join(SCRIPT_DIR, 'sql', 'loc_acct_rollforward.sql')).read()
    mods_sql = open(os.path.join(SCRIPT_DIR, 'sql', 'gwc_mods.sql')).read()

    con = _connect()

    print(f"  Querying LOC tape for {date_end} (EOP)...")
    df_loc = pd.read_sql_query(tape_sql.format(date_end.isoformat()), con)
    print(f"    {len(df_loc)} rows")

    print(f"  Querying rollforward {date_beg} -> {date_end}...")
    df_rollforward = pd.read_sql_query(
        rf_sql.format(date_beg.isoformat(), date_end.isoformat()), con)
    print(f"    {len(df_rollforward)} rows")

    # ── Step 2: Query GWC mods (repayment plans only) ──────────────
    print(f"  Querying GWC mods (RPP%) for {date_end}...")
    df_mods = pd.read_sql_query(mods_sql.format(date_end.isoformat()), con)
    print(f"    {len(df_mods)} rows")

    con.close()

    # ── Step 3: Download template ──────────────────────────────────
    output_dir = os.environ.get('OUTPUTS_PATH', '/mnt/user-data/outputs')
    os.makedirs(output_dir, exist_ok=True)
    template_path = os.path.join(output_dir, '_template_portfolio.xlsx')

    print(f"  Downloading template from Drive (id={args.template_id})...")
    _download_template(args.template_id, template_path)
    print(f"    Template downloaded ({os.path.getsize(template_path) / 1024 / 1024:.1f} MB)")

    # ── Step 4: Shift MoM — copy col L values to col N ─────────────
    # Load with data_only=True to get computed values from col L
    print("  Reading template computed values (data_only)...")
    wb_data = openpyxl.load_workbook(template_path, data_only=True)

    # Collect col L values from both Summary tabs
    summary_values = {}
    for sheet_name in wb_data.sheetnames:
        if sheet_name.startswith('Summary'):
            summary_values[sheet_name] = _get_col_l_values(wb_data[sheet_name])
            print(f"    {sheet_name}: {len(summary_values[sheet_name])} values in col L")
    wb_data.close()

    # Load with data_only=False to preserve formulas
    print("  Loading template with formulas...")
    wb = openpyxl.load_workbook(template_path, data_only=False)

    # Write cached col L values into col N (prior period)
    for sheet_name, values in summary_values.items():
        shifted = _shift_summary_mom_with_values(wb[sheet_name], values)
        print(f"    {sheet_name}: shifted {shifted} values from col L -> col N")

    # ── Step 5: Replace data tabs ──────────────────────────────────
    print("  Replacing data tabs...")

    # loc and rollforward have a pandas index in col A (the original report
    # was written with df.to_excel which includes the index by default).
    # Summary formulas reference columns by letter, so the index column
    # must be present or every formula shifts left by one.
    # mods does NOT have an index column.
    print(f"    loc: {len(df_loc)} rows x {len(df_loc.columns)} cols (+ index)")
    _replace_data_sheet(wb, 'loc', df_loc, include_index=True)

    print(f"    rollforward: {len(df_rollforward)} rows x {len(df_rollforward.columns)} cols (+ index)")
    _replace_data_sheet(wb, 'rollforward', df_rollforward, include_index=True)

    print(f"    mods: {len(df_mods)} rows x {len(df_mods.columns)} cols")
    _replace_data_sheet(wb, 'mods', df_mods, include_index=False)

    # ── Step 6: Save ───────────────────────────────────────────────
    filename = f"Portfolio Report - {month_str}01.xlsx"
    output_path = os.path.join(output_dir, filename)

    print(f"  Saving {filename}...")
    wb.save(output_path)
    wb.close()

    # Clean up template
    try:
        os.remove(template_path)
    except OSError:
        pass

    print(f"\nDone! Output: {output_path}")
    print(f"  LOC tape: {len(df_loc)} accounts")
    print(f"  Rollforward: {len(df_rollforward)} accounts")
    print(f"  Mods (RPP): {len(df_mods)} repayment plans")
    print(f"\nReminder: Open in Excel to recalculate formulas in Summary tabs.")


if __name__ == '__main__':
    main()
