#!/usr/bin/env python3
"""Build the Jeeves Capital Markets Diligence Registry Excel file."""

import os, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

OUTPUTS = os.environ.get('OUTPUTS_PATH', '/mnt/user-data/outputs')
os.makedirs(OUTPUTS, exist_ok=True)

wb = openpyxl.Workbook()

# ─── Color palette ────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F3864")
SUB_FILL   = PatternFill("solid", fgColor="2E75B6")
GRP_FILL   = PatternFill("solid", fgColor="D6E4F7")
EVEN_FILL  = PatternFill("solid", fgColor="F2F7FD")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
GREY_FILL   = PatternFill("solid", fgColor="D9D9D9")

HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
GRP_FONT  = Font(bold=True, color="1F3864", name="Calibri", size=10)
BODY_FONT = Font(name="Calibri", size=9)
BOLD_FONT = Font(bold=True, name="Calibri", size=9)

thin = Side(style='thin', color="BFBFBF")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr_cell(ws, row, col, value, span=1):
    c = ws.cell(row=row, column=col, value=value)
    c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    return c


def grp_cell(ws, row, col, value, span=1):
    c = ws.cell(row=row, column=col, value=value)
    c.font = GRP_FONT; c.fill = GRP_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thin_border
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    return c


def body_cell(ws, row, col, value, fill=None, bold=False, wrap=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = BOLD_FONT if bold else BODY_FONT
    if fill: c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = thin_border
    return c


def status_cell(ws, row, col, status):
    fills = {
        "GREEN": GREEN_FILL, "YELLOW": YELLOW_FILL,
        "RED": RED_FILL, "CLOSED": GREY_FILL, "N/A": WHITE_FILL
    }
    c = ws.cell(row=row, column=col, value=status)
    c.font = BOLD_FONT
    c.fill = fills.get(status, WHITE_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border
    return c


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Master Registry
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Master Registry"
ws.freeze_panes = "A3"

COLS = [
    "Counterparty", "Stage", "Diligence Type", "Item / Document",
    "Status", "Drive File / Folder ID", "Last Updated", "Owner", "Notes"
]

ws.merge_cells("A1:I1")
title = ws["A1"]
title.value = f"Jeeves Capital Markets — Diligence Registry  |  Generated {datetime.today().strftime('%Y-%m-%d')}"
title.font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
title.fill = HDR_FILL
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 22

for i, col in enumerate(COLS, 1):
    hdr_cell(ws, 2, i, col)
ws.row_dimensions[2].height = 28

# (counterparty, stage, dd_type, item, status, drive_id, last_updated, owner, notes)
REGISTRY = [
    # ── BBVA ──
    ("BBVA", "DD + Legal", "Full DD Tracker", "BBVA Due Diligence Tracker (40 items)", "YELLOW", "1CsnL7ftqdSzVuVNQ_k4vgyaMWHslqSMX", "2026-04-20", "Brian", "9 items outstanding per Apr 2026 reconciliation"),
    ("BBVA", "DD + Legal", "Full DD Tracker", "Priority Items Tracker (9 items)", "YELLOW", "15KGEBwaP0S_49fGSXXKac65fnPaBJhQo", "2026-04-08", "Brian", "Subset of most critical open items"),
    ("BBVA", "DD + Legal", "Full DD Tracker", "BBVA Comments File", "GREEN", "1J9A7pEAIEiEEcHikc71p88iEzFYHCovb", "2026-04-07", "Brian", "Jeeves responses to initial BBVA comments"),
    ("BBVA", "DD + Legal", "Full DD Tracker", "Remainder Items Tracker (Google Sheet)", "YELLOW", "1eps6x-lyaBzxJLUp_AYnTsDP3-NI3lxKlbMwVPcylno", "2026-04-16", "Brian", "Live Google Sheet with remainder items"),
    ("BBVA", "DD + Legal", "Full DD Tracker", "Tracker Summary Doc (Word)", "YELLOW", "1p6GuQU0O3PuNwC4usoAtkpNEh5idFPNY", "2026-04-21", "Brian", "Narrative summary of tracker status"),
    ("BBVA", "DD + Legal", "Portfolio", "Loan Tape Folder (3.01)", "GREEN", "1Wt48KhFcG__4SCOeh6yFe8hKnOd6mPL8", "2026-04-16", "Brian", ""),
    ("BBVA", "DD + Legal", "Portfolio", "CAC LTV Metrics Folder (3.03)", "GREEN", "1cMkSmz5R6Yli-h_aaBkEYUxBenCwiSI8", "2026-04-16", "Brian", ""),
    ("BBVA", "DD + Legal", "Portfolio", "Provisioning & Revenue Policy (3.04)", "GREEN", "1hNH1tw11KYiI33STV0fagHSTUiNIKkWg", "2026-04-16", "Brian", ""),
    ("BBVA", "DD + Legal", "Financial", "Jeeves Inc. 2025 Consolidated P&L", "GREEN", "1lWK-Jz6ul3RowaWFZGMznDXSmLS0oeMg", "2026-04-20", "Brian", ""),
    ("BBVA", "DD + Legal", "Financial", "Interim Consolidated Financials (Feb 2026)", "GREEN", "1bH3opQHjveaqOa6dpEhY020s5HnvY2th", "2026-04-20", "Brian", ""),
    ("BBVA", "DD + Legal", "Financial", "Jeeves MX 2024 Financials", "GREEN", "1AvZeA5_Z2Voo58yRT4r_Mr_FYVIi2PXW", "2026-04-20", "Brian", ""),
    ("BBVA", "DD + Legal", "Financial", "Jeeves MX 2025 FS", "GREEN", "1GLXMdDEc70ackQhmTMVmnWANYyngrNrg", "2026-04-28", "Brian", "Latest MX financials"),
    ("BBVA", "DD + Legal", "Financial", "2025 Revenue Summary Breakout", "GREEN", "1VwM63vlBqTVokCFJbqIIOkXqYf1y0buT", "2026-04-23", "Brian", ""),
    ("BBVA", "DD + Legal", "Financial", "Cash Positions Folder (5.05)", "GREEN", "1HVnVyPVJA3dnmohwMsUDZv3qkPxg-030", "2026-04-16", "Brian", ""),
    ("BBVA", "DD + Legal", "Financial", "Bank Statements Folder (5.06)", "GREEN", "1UCWfxvT35qQ0U_-udQtZdHE21cSL2Hqp", "2026-04-16", "Brian", ""),
    ("BBVA", "DD + Legal", "Financial", "Tax Loss Carryforwards Folder (5.07)", "YELLOW", "1vnEZRFcNU8b82E-fua-x-5nOLHOV0OMx", "2026-04-16", "Brian", "Outstanding item #5"),
    ("BBVA", "DD + Legal", "Financial", "Financial Liabilities Folder (5.08)", "YELLOW", "1BPRczn1OnkwgTPJiuzmO4SAmHaChQpr1", "2026-04-16", "Brian", "Outstanding item #6"),
    ("BBVA", "DD + Legal", "Compliance", "Annual Tax Return 2023", "GREEN", "1bEcaKy2HWyaZMrTkgB6Zeo7-hVMMt7q0", "2026-04-28", "Brian", ""),
    ("BBVA", "DD + Legal", "Compliance", "Annual Tax Return 2024", "GREEN", "1FoVR24PZnwp69JeK-hdlbbkihOAFTp34", "2026-04-28", "Brian", ""),
    ("BBVA", "DD + Legal", "Compliance", "Tax Compliance Certificate Folder (7.01)", "YELLOW", "1NvA5US-ijVhf8x3joDHC6i7O_xCrdxwo", "2026-04-16", "Brian", "Outstanding item #8"),
    ("BBVA", "DD + Legal", "Compliance", "Social Security Receipts Folder (7.02)", "YELLOW", "1p8ZX7sTO75bRE7Nw_kOLmZDw9Tb6muqM", "2026-04-16", "Brian", "Outstanding item #9"),
    ("BBVA", "DD + Legal", "Compliance", "Proof of Address Folder (7.03)", "GREEN", "1t4gQXfTizePaqYWr5ZdHE21cSL2Hqp", "2026-04-16", "Brian", ""),
    ("BBVA", "DD + Legal", "Compliance", "Government Permits Folder (7.04)", "GREEN", "1N_GOq1ctXH6FhOyMiEhrZPT7me2d_8eb", "2026-04-16", "Brian", "US, MX, CO confirmed"),
    ("BBVA", "DD + Legal", "Risk/Policy", "Operational Risk Folder (4.01)", "GREEN", "15sPHRNTWWMbg0CP2hXw9Rf2noYXDLOzu", "2026-04-16", "Brian", ""),
    ("BBVA", "DD + Legal", "Risk/Policy", "Hedging Policy Folder (4.02)", "GREEN", "1M5QG7QiYFTtWFN-18_5S3eyKjT8vi2t2", "2026-04-16", "Brian", "FX hedging policy doc — do not describe without reading"),
    ("BBVA", "DD + Legal", "Risk/Policy", "Headcount Folder (4.03)", "GREEN", "1lAd64Om7qjMbOR7cTQsFV4Osdk7Siz1N", "2026-04-16", "Brian", ""),
    ("BBVA", "DD + Legal", "Risk/Policy", "Bank Accounts Folder (4.05)", "GREEN", "11k8etE958y2U1WPH35UFuUVam8b0hTOD", "2026-04-16", "Brian", ""),
    ("BBVA", "DD + Legal", "Risk/Policy", "Key KPIs Folder (4.06)", "GREEN", "1AiKfcXqJsLvLq2xaf7IcqSLq8ZZgNmzr", "2026-04-16", "Brian", ""),
    ("BBVA", "DD + Legal", "Corporate", "Competitive Analysis Folder (1.03)", "YELLOW", "1OZYCAYhC7OogA6SoexVRc7xovuCCKpQT", "2026-04-16", "Brian", "Outstanding item #2"),
    ("BBVA", "DD + Legal", "Corporate", "Org Chart Folder (1.05)", "YELLOW", "10_UnAZDi0Z1-CJSarAOCFZIoKFRUKI3n", "2026-04-16", "Brian", "Outstanding item #3"),
    ("BBVA", "DD + Legal", "Corporate", "Tech Infrastructure Folder (1.07)", "YELLOW", "1B4neiL1rzmQJ57dmy7zzljP72S0kcne5", "2026-04-16", "Brian", "Outstanding item #4"),
    ("BBVA", "DD + Legal", "Corporate", "Bylaws Folder (2.04)", "GREEN", "1ltAv37YgogKjXWh1IwTlq9tlRW9Dm1Gp", "2026-04-16", "Brian", ""),
    ("BBVA", "DD + Legal", "Corporate", "Compliance Folder (2.03)", "RED", "1hTvMN21neYlaRmtHYVUssYN_TWnulPe2", "2026-04-21", "Brian", "Outstanding item #1 — details need confirmation"),
    ("BBVA", "DD + Legal", "Legal", "Master Agreements Folder (6.01)", "YELLOW", "1ixjJOpJnPC01ZZcbANnwMzq8r7svlLQc", "2026-04-16", "Brian", "Outstanding item #7"),
    ("BBVA", "DD + Legal", "Legal", "Collaboration Agreements Folder (6.02)", "GREEN", "11EMRWTyJO5F8_8402KFWFy0GYNqgs9wm", "2026-04-16", "Brian", "US, MX, CO licensing confirmed"),
    ("BBVA", "DD + Legal", "Legal", "Soft TS SBLC + Overdraft + Revolving v4", "GREEN", "1YUbtN2y1O95FK6mTUCCKPaU0iqX7m5Bg", "2026-04-20", "Brian", ""),
    ("BBVA", "DD + Legal", "Legal", "Underwriting Process (Q12)", "GREEN", "1M2s9yq62i4NP8Zsmh7l5ggUpIHf_-dvJ", "2026-04-21", "Brian", ""),
    # ── Neuberger Berman ──
    ("Neuberger Berman", "Active Facility (Signed)", "Legal", "Term Sheet (Executed) — $100MM SOFR+7.5% 24mo", "GREEN", "Debt/Neuberger Berman/Legal/Term Sheet/", "2026-04", "Brian/Goodwin", "Expandable to $150MM. Executed April 2026."),
    ("Neuberger Berman", "Active Facility (Signed)", "Legal", "Collateral Proposal (External)", "GREEN", "1nutFzOzbcsywK2MJG1F96VjGhVhL-Jse", "2026-04-30", "Goodwin/Brian", "Sent to NB"),
    ("Neuberger Berman", "Active Facility (Signed)", "Full DD", "Diligence Session Deck (Apr 20)", "GREEN", "1L7yhWpzX8RDK-AWc_69sTollZcJ_LR_eWxCggHufd9c", "2026-04-21", "Brian", "Presented at NB diligence session"),
    ("Neuberger Berman", "Active Facility (Signed)", "Full DD", "Diligence Session Deck PDF (Apr 22)", "GREEN", "19sKB622S1U2S4Nhx8JuFjcjY1FgC7_Pc", "2026-04-23", "Brian", "Final version shared with NB"),
    ("Neuberger Berman", "Active Facility (Signed)", "Full DD", "Diligence Agenda (Apr 20)", "GREEN", "123tNk1hU9tkiJ5yGhgY0EKKPvi8FCmSY", "2026-04-20", "Brian", ""),
    ("Neuberger Berman", "Active Facility (Signed)", "Background", "Background Check Authorization (Dileep)", "GREEN", "1IzDJhCCK7UwG00BJgrkYCDVikxqDreSS", "2026-04-21", "Brian", ""),
    ("Neuberger Berman", "Active Facility (Signed)", "Portfolio", "LOC Account Rollforward (Jan 2024 - Mar 2026)", "GREEN", "1xSDMKThEnERR4SEKOrKP4z8sipfJV9GV", "2026-04-29", "Brian", "17.8MB — uses loc_acct_rollforward.sql"),
    ("Neuberger Berman", "Active Facility (Signed)", "Financial", "Prelim Financial Forecast (May 5)", "GREEN", "1YKr-YWYXItN1DNOyyY5b4VfkefDxZoga", "2026-05-05", "Brian", ""),
    ("Neuberger Berman", "Active Facility (Signed)", "Financial", "Prelim Financial Forecast (May 6 — latest)", "GREEN", "1bAMVUeG5SHFqVuNw1cMKHc3z1eLUD4bR", "2026-05-06", "Brian", "Latest version"),
    ("Neuberger Berman", "Active Facility (Signed)", "Analysis", "NB vs CIM (Bridge + CO SPV) Analysis", "GREEN", "1Zh4TmLJYakvp9UP8TG_rDOAV7lE8DBCm", "2026-04-30", "Brian", "Decision memo: NB=Brazil, CIM=Colombia"),
    # ── Francisco Partners ──
    ("Francisco Partners", "Diligence", "Portfolio", "Monthly Snapshot Tapes (through May 2026)", "GREEN", "1nmXo4zLkeBLUT3SdhmjX3nMcfJbTMhMy", "2026-05-07", "Brian", "CO filter removed; charge_off_flag=TRUE included. 61.5MB."),
    ("Francisco Partners", "Diligence", "Portfolio", "Portfolio Rollforward", "GREEN", "1NVO0bs8g_MQl9pSo1H08mz06XPLLzgaJ", "2026-05-07", "Brian", "Uses loc_acct_rollforward.sql"),
    ("Francisco Partners", "Diligence", "Portfolio", "Top 20 Borrowers", "GREEN", "1Ub0YQgf2uHOqdA43mQ9eN8Ds13j0r3ST", "2026-05-07", "Brian", ""),
    ("Francisco Partners", "Diligence", "Overview", "Facility Overview / VDR Summary", "GREEN", "1hv3KJO5IwiEM1q5fC20gZNwJ2D-EH8Dh", "2026-05-08", "Brian", "4 files in Diligence folder — complete as of May 2026"),
    # ── Vista Credit ──
    ("Vista Credit", "DDQ Completed", "DDQ", "DDQ List from VCP (original)", "GREEN", "1_IQTBpW0PG841L81qCskuHKg9_GWGTEy", "2026-04-17", "Vista Credit", "PDF of questions from counterparty"),
    ("Vista Credit", "DDQ Completed", "DDQ", "DDQ Response Document", "GREEN", "1VqHwoxs4RxVMJ_Up9DsM63XjhyIZIw0jC_UL5O68M40", "2026-04-20", "Brian", "Completed Google Doc response"),
    ("Vista Credit", "DDQ Completed", "DDQ", "DDQ Tracker (Google Sheet)", "GREEN", "1j7L7Hz0oOipaFMu14c2vgblIMje9eWWVlO4LaufVVKA", "2026-04-17", "Brian", ""),
    ("Vista Credit", "DDQ Completed", "Portfolio", "Data Package for DDQ", "GREEN", "1U6xRs0ZONhVhVgl0S8rvyLiyRNYhbE6jmFPVa_urB4U", "2026-04-18", "Brian", "Google Sheet portfolio data"),
    # ── CIM ──
    ("CIM", "Active Facility", "Legal", "Colombia SPV Legal Folder", "YELLOW", "1LIDqvnRVcetsg8BiWQ6tOSOn5ov9XWFj", "2026-04-30", "Goodwin/Brian", "5th Amendment in progress. CIM pushing CO SPV."),
    ("CIM", "Active Facility", "Diligence", "CO SPV Diligence Folder", "YELLOW", "1PgPNdfXRd3i-1dI09sp_mx0saP47ajsy", "2026-02-10", "Brian", "CIM CO SPV diligence materials"),
    ("CIM", "Active Facility", "Diligence", "Corp Credit Diligence Folder", "GREEN", "1Dnl-BZje2n0ffZvOGbOOBmNp6fLlO1bH", "2026-01-30", "Brian", ""),
    ("CIM", "Active Facility", "Diligence", "MX SPV Diligence Folder", "GREEN", "1DZbxyD4Rt9tiWUt5RYvqKTuMFf2Aj97W", "2026-01-30", "Brian", ""),
    ("CIM", "Active Facility", "Analysis", "CIM vs NB Colombia Term Sheet Comparison", "GREEN", "1CcKkBtZEh0NPnysIF3_uxbmQ1hPG67Jq", "2026-04-10", "Brian", ""),
    ("CIM", "Active Facility", "Analysis", "NB vs CIM CO SPV Analysis (Colombia)", "GREEN", "1ZaILll_piV8vUveDMCueWms1IDsyOLx2", "2026-04-30", "Brian", ""),
    ("CIM", "Active Facility", "Audit", "Jan 2026 Audit Folder", "GREEN", "1RezTWe_zv99wBk6IklRcR8aGP1AJNmsT", "2026-01-12", "Brian", ""),
    # ── Covalto ──
    ("Covalto", "Term Sheet Negotiation", "Legal", "Secured Term Sheet (April 2026)", "GREEN", "1b7-okWAiCl8NrF1F_PGq-1Xr4RTgKcCU", "2026-04-08", "Brian", "Secured facility TS in Spanish + English"),
    ("Covalto", "Term Sheet Negotiation", "Legal", "Redline — Covalto TS (Mar 30)", "GREEN", "1iiY-b9q6jdTszXIKVz10GuCJmn7aIpLa", "2026-03-30", "Brian", ""),
    ("Covalto", "Term Sheet Negotiation", "Legal", "Redline — Covalto TS (Apr 3)", "GREEN", "1m3d8R7GgHG737AXjVPozGWLKCYLL--Tv", "2026-04-03", "Brian", "Most recent redline round"),
    ("Covalto", "Term Sheet Negotiation", "Legal", "Jeeves Comments (Apr 3)", "GREEN", "1qWlula9PjCQXLUNreCtkLMCg7zo7x_kO", "2026-04-03", "Brian", ""),
    ("Covalto", "Term Sheet Negotiation", "Legal", "Jeeves Comments (Apr 4 — latest)", "GREEN", "1Z3HMBlxoHXYeAPpCV2RVAfmC9ErHBfHE", "2026-04-04", "Brian", "Most recent Jeeves position"),
    ("Covalto", "Term Sheet Negotiation", "Portfolio", "MC Data Analysis", "GREEN", "1el26DfbfgcS47iF5rxhIw-FuOUJ1l7aV", "2026-04-17", "Brian", ""),
    # ── Gramercy ──
    ("Gramercy", "Term Sheet Negotiation", "Legal", "Corp Facility Term Sheet (Form)", "GREEN", "1Ssiqj_98LoE9jjJvTjzXLe7uLwua2cLg", "2026-03-18", "Brian", ""),
    ("Gramercy", "Term Sheet Negotiation", "Legal", "Gramercy Comments (Mar 20)", "GREEN", "1R2-MOBP5qpI120MdETfGxa9kKii4FC7M", "2026-03-25", "Brian", ""),
    ("Gramercy", "Term Sheet Negotiation", "Legal", "Gramercy Comments (Apr 1)", "GREEN", "1mkDpIIdjIBvTWn1yrZYOO3eVbedaFJKP", "2026-04-01", "Brian", ""),
    ("Gramercy", "Term Sheet Negotiation", "Legal", "Jeeves Comments (Mar 26)", "GREEN", "1sYqWea_p1nQdqi5KA7UgXSkraGXTIjFt", "2026-03-26", "Brian", ""),
    ("Gramercy", "Term Sheet Negotiation", "Legal", "Jeeves Comments (Mar 31)", "GREEN", "1aO1E2qmp99h-YKndYYutNRyEQ_ghhoIS", "2026-03-31", "Brian", ""),
    ("Gramercy", "Term Sheet Negotiation", "Legal", "Redline — Gramercy TS (Apr 1)", "GREEN", "1LVENcdL55VQXeVQ7ID15OZbR2AZW-QL4", "2026-04-01", "Brian", "Most recent round"),
    # ── Fasanara ──
    ("Fasanara", "Term Sheet Negotiation", "Legal", "Term Sheet v1 (Mar 26)", "GREEN", "1c4MQiV0qh8e892tw9XH13jnwQM-NXUm8", "2026-03-28", "Brian", ""),
    ("Fasanara", "Term Sheet Negotiation", "Legal", "Jeeves Comments (Mar 26)", "GREEN", "1UkBC8aLkxsMFXXkC67p46KBkjHHiLIHq", "2026-03-31", "Brian", ""),
    ("Fasanara", "Term Sheet Negotiation", "Legal", "Jeeves Comments (Mar 31)", "GREEN", "1LOSStgnPspWEsLXvorPiKY15Vg_ag_L0", "2026-03-31", "Brian", ""),
    ("Fasanara", "Term Sheet Negotiation", "Legal", "Term Sheet v2 (Apr 2)", "GREEN", "16jjBHo3Fut1xGoUcOHWKyEq4irS-nAH_", "2026-04-02", "Brian", ""),
    ("Fasanara", "Term Sheet Negotiation", "Legal", "Redline — Fasanara Terms v2 vs Jeeves Comments (Apr 2)", "GREEN", "1IPkw05C16T8QraOJL-kMvgUkb4ksOyDu", "2026-04-02", "Brian", "Most recent round"),
    ("Fasanara", "Term Sheet Negotiation", "DDQ", "Jeeves Q&A Spreadsheet", "GREEN", "1S3GCM4TpjlsvBLPhWaD2WHYupKLtQ_iZ", "2026-03-26", "Brian", ""),
    # ── i80 ──
    ("i80", "Data Room", "Portfolio", "Data Room Package (Feb 2026)", "GREEN", "1EhSsaTBijTp397slsGxeZ2cmBy1kFNXr", "2026-02-25", "Brian", "Corporate deck, data tape, financials shared"),
    # ── Accial ──
    ("Accial", "Term Sheet", "Legal", "Term Sheet (Mar 2026)", "GREEN", "1TQQYTyMZju_SnRaO7jgSFidIgeKYhyva", "2026-03-25", "Brian", ""),
    # ── BTG AM ──
    ("BTG AM", "Q&A Stage", "DDQ", "FIDC Q&A Response Document", "GREEN", "1C5hP_2RLsY4MCBSUNeIU8", "2026-01-16", "Brian", "Responded to FIDC Q&A in Jan 2026"),
    ("BTG AM", "Q&A Stage", "Legal", "NDA", "GREEN", "1BHXPJLgNw_erAJoCMBlnCUPrV", "2026-01-14", "Brian", ""),
    # ── Lendable / Rivonia / PFG / UBS ──
    ("Lendable", "NDA Only", "Legal", "NDA", "CLOSED", "19qgV99WF-AnpMSgbtPbKNKof46fKxtt0", "2026-02", "Brian", "Feb 2026"),
    ("Rivonia Road", "NDA Only", "Legal", "NDA (Signed)", "CLOSED", "1gZIxELrnVNbIiQeNfbqsoakXJPiXx_BI", "2026-03-16", "Brian", "Mar 2026"),
    ("PFG", "NDA Only", "Legal", "NDA Folder", "CLOSED", "1W6BIbJvU_vVIvY_XhtklIFx337TsyZWG", "2026-02", "Brian", "Feb 2026"),
    ("UBS", "NDA + Discussion", "Legal", "NDA (Jan 2025)", "CLOSED", "1AAQBlQydpSpAy179Oj0HPIelWn2LdrWK", "2025-01-08", "Brian", "Discussions ongoing Jan-Mar 2025"),
    # ── Cross-counterparty ──
    ("[Cross-CP]", "Analysis", "Analysis", "Lender Term Sheet Comparison (Apr 1)", "GREEN", "1QKtFiEpIBCy5GKWh55Bfs34avGds8N9z", "2026-04-02", "Brian", "All active term sheets vs CIM"),
    ("[Cross-CP]", "Analysis", "Analysis", "CIM vs NB Colombia Term Sheet Comparison", "GREEN", "1CcKkBtZEh0NPnysIF3_uxbmQ1hPG67Jq", "2026-04-10", "Brian", ""),
    ("[Cross-CP]", "Analysis", "Analysis", "Facility Overview VDR (Apr 23)", "GREEN", "1ZNzPfwZxHQrxnoCLtwDl_bjqnMQr415U", "2026-04-24", "Brian", "Master facility summary across all lenders"),
    ("[Cross-CP]", "Analysis", "Analysis", "Lender Term Sheet Comparison Doc (Apr 2)", "GREEN", "13PZSY_P_saxsNDqRpnIfMTCg7PjQjFJa", "2026-04-02", "Brian", ""),
]

row = 3
current_cp = None

for entry in REGISTRY:
    cp, stage, dd_type, item, status, drive_id, updated, owner, notes = entry
    fill = EVEN_FILL if row % 2 == 0 else WHITE_FILL

    if cp != current_cp:
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = GRP_FILL
            ws.cell(row=row, column=col).border = thin_border
        grp_cell(ws, row, 1, f"  {cp}  |  {stage}", span=9)
        ws.row_dimensions[row].height = 16
        row += 1
        current_cp = cp

    body_cell(ws, row, 1, cp, fill=fill, bold=True)
    body_cell(ws, row, 2, stage, fill=fill)
    body_cell(ws, row, 3, dd_type, fill=fill)
    body_cell(ws, row, 4, item, fill=fill, wrap=True)
    status_cell(ws, row, 5, status)

    c = ws.cell(row=row, column=6, value=drive_id)
    c.font = Font(name="Calibri", size=9, color="1F497D", underline="single")
    c.fill = fill; c.border = thin_border
    c.alignment = Alignment(horizontal="left", vertical="center")
    if len(drive_id) > 20 and not drive_id.startswith("Debt/"):
        c.hyperlink = f"https://drive.google.com/file/d/{drive_id}/view"

    body_cell(ws, row, 7, updated, fill=fill, align="center")
    body_cell(ws, row, 8, owner, fill=fill, align="center")
    body_cell(ws, row, 9, notes, fill=fill, wrap=True)
    ws.row_dimensions[row].height = 30
    row += 1

col_widths = [20, 24, 18, 44, 10, 38, 13, 14, 42]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Counterparty Summary
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Counterparty Summary")
ws3.freeze_panes = "A3"

ws3.merge_cells("A1:H1")
t = ws3["A1"]
t.value = "Counterparty Status Summary  |  Last 6 Months  |  As of 2026-05-07"
t.font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
t.fill = HDR_FILL
t.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 22

cols3 = ["Counterparty", "Stage", "Product / Facility", "Size", "Rate", "Status", "Drive Folder", "Key Contacts / Counsel"]
for i, c in enumerate(cols3, 1):
    hdr_cell(ws3, 2, i, c)
ws3.row_dimensions[2].height = 28

summary_data = [
    ("BBVA",             "DD + Legal (Active)",          "SBLC + Overdraft + Revolving",   "TBD",             "TBD",        "YELLOW", "12ns4FGnFiA6K3jH3h6cECJ2S8TD8irEf", "White & Case"),
    ("Neuberger Berman", "Facility Signed (Apr 2026)",   "Colombia LOC SPV",               "$100MM (exp $150MM)", "SOFR+7.5%", "GREEN",  "1P6nwDSUK5MnoX7FLEfEckcGwqCgsWRbx", "Goodwin, BU Colombia"),
    ("Francisco Partners","Diligence",                   "TBD",                            "TBD",             "TBD",        "YELLOW", "1LdmMpCmQQ5Y1UUDoxNnAZ1toWIrytJp4", "FP team"),
    ("Vista Credit",     "DDQ Completed",                "TBD",                            "TBD",             "TBD",        "GREEN",  "1ah1x2cD_wIBQrRku7xuLelS52-D0L3I8", "Vista CP team"),
    ("CIM",              "Active Facility",              "Bridge LOC (US + MX); CO SPV",   "Active",          "Contract",   "YELLOW", "1bdplsyvIQPtVKAMe6Z_XzeceJYxFKwFB", "Alejandra Granados, Goodwin"),
    ("Covalto",          "Term Sheet Negotiation",       "Secured Facility (MX)",          "TBD",             "TBD",        "YELLOW", "11v7G67k_XSGVXn7igUTRJlVNeojmcpZO", "Covalto team"),
    ("Gramercy",         "Term Sheet Negotiation",       "Corp Facility",                  "TBD",             "TBD",        "YELLOW", "1k-R1fldUnw90kZpJCS7VR5Yu7SNu0TXn", "Gramercy team"),
    ("Fasanara",         "Term Sheet Negotiation",       "Credit Facility",                "TBD",             "TBD",        "YELLOW", "125_p3cKygzuyh-dbcarZjMP9HI74ohhx", "Fasanara team"),
    ("i80",              "Data Room",                    "TBD",                            "TBD",             "TBD",        "YELLOW", "1EhSsaTBijTp397slsGxeZ2cmBy1kFNXr", "i80 team"),
    ("Accial",           "Term Sheet",                   "TBD",                            "TBD",             "TBD",        "YELLOW", "1A_5Ux5v1l04NypmpsXL4ChE1sNmYcx1R", "Accial team"),
    ("BTG AM",           "Q&A Stage",                    "FIDC Brazil",                    "TBD",             "TBD",        "YELLOW", "1B8e9rX-Wqaz_qX7DxHtXzh2ZtGvz3uRC", "BTG AM team"),
    ("Lendable",         "NDA Only",                     "TBD",                            "—",               "—",          "CLOSED", "19qgV99WF-AnpMSgbtPbKNKof46fKxtt0", "Lendable team"),
    ("Rivonia Road",     "NDA Only",                     "TBD",                            "—",               "—",          "CLOSED", "1gwZtF1lnCklkU20JJgMD0sILDO-XMN8z", "Rivonia Road team"),
    ("PFG",              "NDA Only",                     "TBD",                            "—",               "—",          "CLOSED", "1W6BIbJvU_vVIvY_XhtklIFx337TsyZWG", "PFG team"),
    ("UBS",              "NDA + Discussion (2025)",      "TBD",                            "—",               "—",          "CLOSED", "1A2ipoW1WaRxI9Hbgeloq6SAqKXEMgOp-", "UBS team"),
]

status_fills = {
    "GREEN": GREEN_FILL, "YELLOW": YELLOW_FILL,
    "RED": RED_FILL, "CLOSED": GREY_FILL
}

for i, row_data in enumerate(summary_data, 3):
    cp, stage, product, size, rate, status, folder_id, contacts = row_data
    fill = EVEN_FILL if i % 2 == 0 else WHITE_FILL
    body_cell(ws3, i, 1, cp, fill=fill, bold=True)
    body_cell(ws3, i, 2, stage, fill=fill)
    body_cell(ws3, i, 3, product, fill=fill)
    body_cell(ws3, i, 4, size, fill=fill, align="center")
    body_cell(ws3, i, 5, rate, fill=fill, align="center")
    s = ws3.cell(row=i, column=6, value=status)
    s.font = BOLD_FONT; s.fill = status_fills.get(status, WHITE_FILL)
    s.alignment = Alignment(horizontal="center", vertical="center"); s.border = thin_border
    c = ws3.cell(row=i, column=7, value=folder_id)
    c.hyperlink = f"https://drive.google.com/drive/folders/{folder_id}"
    c.font = Font(name="Calibri", size=9, color="1F497D", underline="single")
    c.fill = fill; c.border = thin_border
    c.alignment = Alignment(horizontal="left", vertical="center")
    body_cell(ws3, i, 8, contacts, fill=fill)
    ws3.row_dimensions[i].height = 22

for i, w in enumerate([22, 28, 30, 20, 12, 10, 36, 30], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Monthly Update Runbook
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Monthly Update Runbook")

ws4.merge_cells("A1:C1")
t = ws4["A1"]
t.value = "Monthly Diligence Registry Update — Runbook"
t.font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
t.fill = HDR_FILL
t.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 22

guide_cols = ["Step", "Action", "Notes"]
for i, c in enumerate(guide_cols, 1):
    hdr_cell(ws4, 2, i, c)
ws4.row_dimensions[2].height = 24

guide_rows = [
    ("1", "Run: python diligence_tool.py gather-portfolio --date YYYY-MM-DD",
     "Pull latest portfolio snapshot. Used to verify any portfolio-level DD claims."),
    ("2", "Browse each active counterparty Drive folder for new files since last month run date",
     "Use list_drive_folder.py on: BBVA (1pA5_GOqtHMTatJE5vIIYCwm-p742d5yT), NB (19fmtr7f3714EGe9j8fYFBUHmZ7_aWRz0), FP (1Z82iHprfIyXKdxNeuvwMUSiYXeOCH67X), Vista Credit (1ah1x2cD_wIBQrRku7xuLelS52-D0L3I8), CIM Legal (1LIDqvnRVcetsg8BiWQ6tOSOn5ov9XWFj)"),
    ("3", "Add any new documents to Master Registry sheet with correct status, Drive ID, and date",
     "Only add confirmed files — no placeholders. Every row must have a real Drive ID."),
    ("4", "Update BBVA Outstanding: verify which of the 9 open items closed. Change status GREEN when confirmed.",
     "Outstanding items: 2.03 (RED), 1.03, 1.05, 1.07, 5.07, 5.08, 6.01, 7.01, 7.02 (all YELLOW)"),
    ("5", "Update Counterparty Summary: reflect any stage changes",
     "E.g., Covalto TS Negotiation -> Signed; FP Diligence -> Full DD; NB -> First Draw"),
    ("6", "Re-upload this file to Drive (Debt/ root folder ID: 1-0K8EM8slr1_I4Iik7_ZZn0t4SSAMKLU)",
     "Rename with new date: Diligence Registry - Capital Markets - YYYYMMDD.xlsx\npython upload_to_drive.py 'Diligence Registry - Capital Markets - YYYYMMDD.xlsx' --folder 1-0K8EM8slr1_I4Iik7_ZZn0t4SSAMKLU"),
    ("7", "Patch jeeves-diligence SKILL.md if any new lessons or counterparty facts emerged",
     "python /mnt/skills/public/self-improving-agent/scripts/skill_manage.py log jeeves-diligence --episode '{...}'"),
    ("8", "Optional: Post Slack summary of changes to #capital-markets",
     "Format: bullet list of new items added, items closed GREEN, stage changes"),
]

for i, (step, action, notes) in enumerate(guide_rows, 3):
    fill = EVEN_FILL if i % 2 == 0 else WHITE_FILL
    for col, val, wid in [(1, step, False), (2, action, True), (3, notes, True)]:
        c = ws4.cell(row=i, column=col, value=val)
        c.font = BODY_FONT; c.fill = fill; c.border = thin_border
        c.alignment = Alignment(horizontal="left" if col > 1 else "center", vertical="top", wrap_text=wid)
    ws4.row_dimensions[i].height = 55

ws4.column_dimensions["A"].width = 7
ws4.column_dimensions["B"].width = 60
ws4.column_dimensions["C"].width = 52

# Tab colors
ws.sheet_properties.tabColor   = "1F3864"
ws3.sheet_properties.tabColor  = "2E75B6"
ws4.sheet_properties.tabColor  = "375623"

out = os.path.join(OUTPUTS, "Diligence Registry - Capital Markets - 20260507.xlsx")
wb.save(out)
print(f"Saved: {out}")
